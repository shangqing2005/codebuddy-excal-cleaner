# -*- coding: utf-8 -*-
"""
=============================================================================
  Excel数据清洗工具 v0.4
  功能：跨Sheet单元格去重 - 按指定列范围的内容进行跨Sheet去重，
        删除重复行的单元格内容并上移，支持单规则/多规则审查模式
=============================================================================

  作者：zyq
  联系：13166954800
  日期：2026-03-30
  版本：0.4

---------------------------------------------------------------------------
  运行环境
---------------------------------------------------------------------------
  Python    : 3.10 ~ 3.12（推荐 3.12）
  第三方依赖 :
    - customtkinter >= 5.2   (GUI 框架)
    - openpyxl      >= 3.1   (Excel 读写)
  标准库依赖 :
    - tkinter, threading, time, os, copy, gc, shutil, collections

---------------------------------------------------------------------------
  核心架构
---------------------------------------------------------------------------
  主类        : ExcelDeduplicationTool（基于 CustomTkinter 单窗口应用）
  处理模式    : 流式两遍扫描 + 就地修改副本（低内存占用）
                - 第一遍：只读模式扫描，记录重复行号
                - 第二遍：写模式打开副本，按行号删除重复内容并上移
  并发模型    : 子线程处理文件，主线程更新 GUI（threading + Event）

---------------------------------------------------------------------------
  主要功能模块
---------------------------------------------------------------------------
  文件管理
    - _select_file()             选择 Excel 文件（.xlsx）
    - _show_large_file_dialog()  大文件（>50MB）提醒弹窗，含加载耗时显示与倒计时
    - _load_file_and_init_ui()   加载文件并初始化界面
    - _cleanup_file_state()      切换/关闭文件时清理内存与状态

  Sheet 与列选择
    - _read_headers()            流式读取表头（缓存只读 workbook 加速切换）
    - _close_ro_wb()             释放缓存的只读 workbook
    - _on_sheet_changed()        Sheet 切换事件
    - _on_column_click()         列点击选择（绿色整体范围 / 红色审查列）
    - _redraw_column_cells()     重绘列单元格显示
    - _update_column_display()   更新列区域滚动与布局

  规则管理（单规则 / 多规则）
    - _on_rule_mode_changed()    规则模式切换
    - _save_rule()               保存当前列选择为规则
    - _edit_rule()               编辑已有规则
    - _delete_rule()             删除规则
    - _clear_all_rules()         清空所有规则
    - _refresh_rules_list()      刷新规则列表显示

  数据处理核心
    - _start_processing()        启动处理（检查规则 → 创建副本 → 流式处理）
    - _process_excel_with_rules() 核心去重逻辑：
        · 按规则遍历每个 Sheet
        · 统计各单元格内容出现次数，标记重复项
        · 红色审查列：仅当前 Sheet 内去重（首次出现保留）
        · 绿色整体范围：跨 Sheet 去重（首次出现保留，后续删除）
        · 删除方式：清空重复行在指定列的内容，下方单元格上移填补
    - _toggle_pause()            暂停 / 继续
    - _cancel_processing()       取消处理（清理临时文件）

  计时与日志
    - _start_timer() / _tick_timer() / _stop_timer()  处理耗时计时
    - _log()                     日志输出到界面

---------------------------------------------------------------------------
  列选择说明
---------------------------------------------------------------------------
  绿色（整体范围）: 指定列范围参与跨 Sheet 去重，内容在整个文件中全局唯一
  红色（审查列）  : 单规则模式下额外指定审查列，仅当前 Sheet 内去重

---------------------------------------------------------------------------
  输出说明
---------------------------------------------------------------------------
  输出文件    : 与源文件同目录，命名为 "<原文件名>_已去重.xlsx"
  副本机制    : 使用 shutil.copy2 创建工作副本，处理后另存，不修改原文件
  内存优化    : read_only 流式读取 + gc.collect() 主动回收，适合大文件处理

---------------------------------------------------------------------------
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import threading
import time
import os
import sys
import copy
import gc
import shutil
import json
import subprocess
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from zipfile import BadZipFile
import re


# ==================== 数据归一化：清除隐藏/不可见字符，确保跨Sheet去重一致 ====================
_INVISIBLE_CHARS_RE = re.compile(r'[\ufeff\u200b\u200c\u200d]')  # BOM、零宽字符
_MULTI_SPACE_RE = re.compile(r'\s+')


def normalize_value(raw: str) -> str:
    """清洗单元格值：去除BOM、零宽字符、不间断空格等不可见字符，压缩空白"""
    if not raw:
        return ""
    s = _INVISIBLE_CHARS_RE.sub('', raw)           # 去除BOM和零宽字符
    s = s.replace('\xa0', ' ').replace('\u3000', ' ')  # NBSP、全角空格→普通空格
    s = s.strip()                                  # 去首尾空白
    s = _MULTI_SPACE_RE.sub(' ', s).strip()        # 压缩中间连续空白
    return s


# ==================== 兼容性导入：openpyxl 不同版本的异常类位置不同 ====================
try:
    from openpyxl.utils.exceptions import InvalidFileException
except (ImportError, ModuleNotFoundError):
    try:
        from openpyxl.errors import InvalidFileException
    except (ImportError, ModuleNotFoundError):
        # 兜底：定义一个空基类（理论上不会走到这里）
        InvalidFileException = ValueError


# ==================== 内存优化配置 ====================
# 大文件阈值（MB），超过此值启用内存保护模式
LARGE_FILE_THRESHOLD_MB = 100
# 单次处理的重复项上限（超过时分批写入）
BATCH_DUPLICATE_LIMIT = 50000
# 汇总统计最大保留条数（超出时丢弃低频项）
MAX_SUMMARY_ENTRIES = 5000


# ==================== 自定义异常类 ====================
# 所有业务异常统一继承自基类，便于集中处理

class ExcelCleanerError(Exception):
    """Excel清洗工具基础异常类"""
    def __init__(self, message, error_code=None, suggestion=None):
        self.message = message
        self.error_code = error_code or "UNKNOWN"
        self.suggestion = suggestion
        super().__init__(self.message)

    def get_user_message(self):
        """返回面向用户的友好错误消息（用于弹窗显示）"""
        msg = f"【{self.error_code}】{self.message}"
        if self.suggestion:
            msg += f"\n\n{self.suggestion}"
        return msg


class FileNotFoundError_(ExcelCleanerError):
    """E001 - 文件不存在或路径无效"""
    def __init__(self, file_path):
        super().__init__(
            message=f"文件不存在或路径无效：{file_path}",
            error_code="E001",
            suggestion="请检查文件路径是否正确，确认文件是否已被移动、删除或重命名"
        )


class FilePermissionError_(ExcelCleanerError):
    """E002 - 文件被占用或无访问权限"""
    def __init__(self, file_path, is_locked=False):
        if is_locked:
            msg = f"文件正在被其他程序占用：{file_path}"
            hint = "请关闭Excel、WPS或其他正在打开该文件的程序后重试"
        else:
            msg = f"没有权限访问该文件：{file_path}"
            hint = "请右键点击文件→属性→安全，确保当前账户有读取/写入权限"
        super().__init__(message=msg, error_code="E002", suggestion=hint)


class FileCorruptError(ExcelCleanerError):
    """E003 - Excel文件损坏或格式不兼容"""
    def __init__(self, file_path, detail=None):
        msg = f"Excel文件损坏、格式错误或不兼容：{os.path.basename(file_path)}"
        if detail:
            msg += f"\n技术详情：{detail}"
        super().__init__(
            message=msg,
            error_code="E003",
            suggestion="请尝试以下方法：\n1. 用Excel打开该文件并另存为.xlsx格式\n2. 确认文件未损坏（非下载中断或传输错误）\n3. 确认文件确实是xlsx格式而非xls/csv"
        )


class EmptyFileError(ExcelCleanerError):
    """E004 - Excel文件为空或无可用数据"""
    def __init__(self, file_path):
        super().__init__(
            message=f"Excel文件为空或没有可用的数据：{os.path.basename(file_path)}",
            error_code="E004",
            suggestion="请确认文件中包含至少一个工作表，且工作表内有数据"
        )


class SheetNotFoundError(ExcelCleanerError):
    """E005 - 指定的Sheet工作表不存在"""
    def __init__(self, sheet_name, available_sheets=None):
        msg = f"指定的工作表「{sheet_name}」不存在"
        if available_sheets:
            msg += f"\n当前可用的工作表：{', '.join(available_sheets[:10])}"
            if len(available_sheets) > 10:
                msg += f" ... 共{len(available_sheets)}个"
        super().__init__(
            message=msg,
            error_code="E005",
            suggestion="请检查规则配置中的工作表名称是否正确（注意区分大小写）"
        )


class ColumnOutOfRangeError(ExcelCleanerError):
    """E006 - 列号超出数据范围"""
    def __init__(self, col_idx, max_col, sheet_name=None):
        col_letter = get_column_letter(col_idx + 1)
        max_letter = get_column_letter(max_col) if max_col > 0 else "0"
        info = f"（工作表：{sheet_name}）" if sheet_name else ""
        super().__init__(
            message=f"列索引越界{info}：选择了第{col_letter}列，但该表最大只有{max_letter}列",
            error_code="E006",
            suggestion="请重新选择列范围，确保所选列在数据范围内"
        )


class DiskSpaceError(ExcelCleanerError):
    """E007 - 磁盘空间不足"""
    def __init__(self, required_mb, available_mb):
        super().__init__(
            message=f"磁盘剩余空间不足！需要约 {required_mb:.1f} MB，但当前仅剩 {available_mb:.1f} MB",
            error_code="E007",
            suggestion="请清理磁盘空间后再试：\n1. 清空回收站\n2. 删除不需要的大文件\n3. 或将文件移动到空间充足的磁盘"
        )


class MemoryError_(ExcelCleanerError):
    """E008 - 内存不足，无法完成处理"""
    def __init__(self, file_size_mb=None):
        extra = f"（当前文件约 {file_size_mb:.0f} MB）" if file_size_mb else ""
        super().__init__(
            message=f"系统内存不足，无法处理该文件{extra}",
            error_code="E008",
            suggestion="建议方案：\n1. 关闭其他程序释放内存\n2. 将大文件拆分成多个小文件分别处理\n3. 增加电脑物理内存"
        )


class InvalidRuleError(ExcelCleanerError):
    """E009 - 规则配置无效或缺失"""
    def __init__(self, rule_idx=None, reason=""):
        idx_info = f"第 {rule_idx + 1} 条规则" if rule_idx is not None else "规则"
        super().__init__(
            message=f"{idx_info}配置无效：{reason}",
            error_code="E009",
            suggestion="请检查每条规则是否同时包含：\n1. 整体范围（绿色选择的列）\n2. 审查列（红色选择的列）"
        )


class WriteFileError(ExcelCleanerError):
    """E010 - 写入/保存文件失败"""
    def __init__(self, file_path, operation="保存", detail=None):
        msg = f"{operation}文件失败：{file_path}"
        if detail:
            msg += f"\n原因：{detail}"
        super().__init__(
            message=msg,
            error_code="E010",
            suggestion="请检查：\n1. 磁盘是否有足够的写入空间\n2. 是否有写入该目录的权限\n3. 杀毒软件是否阻止了文件写入"
        )


class UserCancelledError(ExcelCleanerError):
    """E011 - 用户主动取消操作（正常中断，不算错误）"""
    def __init__(self):
        super().__init__(
            message="操作已被用户取消",
            error_code="E011",
            suggestion=None
        )

# 设置 CustomTkinter 外观
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

# 颜色常量
COLOR_BG = "#F8F9FA"                    # 极浅灰背景
COLOR_BG_WHITE = "#FFFFFF"              # 纯白背景
COLOR_BORDER = "#DEE2E6"                # 浅灰蓝边框
COLOR_TEXT_PRIMARY = "#212529"           # 深灰黑主文字
COLOR_TEXT_MUTED = "#6C757D"            # 中灰辅助文字
COLOR_GREEN = "#2E7D32"
COLOR_GREEN_LIGHT = "#A5D6A7"
COLOR_RED = "#C62828"
COLOR_RED_LIGHT = "#EF9A9A"
COLOR_BLUE = "#1565C0"
COLOR_BLUE_LIGHT = "#90CAF9"
COLOR_ORANGE = "#FF9800"
COLOR_YELLOW = "#F57F17"

# 全局字体
FONT_FAMILY = "Microsoft YaHei"
FONT_MONO = "Consolas"
FONT_SIZE_BODY = 16       # 正文 16px
FONT_SIZE_SMALL = 14      # 辅助 14px
FONT_SIZE_TITLE = 18      # 标题 18px
FONT_SIZE_BTN = 14        # 按钮文字 14px


class AutoSelectDialog:
    """自动选择弹窗：跨Sheet自动生成去重规则"""

    def __init__(self, parent, file_path, all_headers_cache):
        self.parent = parent
        self.file_path = file_path
        self.all_headers_cache = all_headers_cache  # {sheet_name: tuple(headers)}
        self.sheet_names = list(all_headers_cache.keys())

        # 选择状态
        self.green_cols = set()      # 当前选中的整体范围列索引(0-based)
        self.red_col = None          # 当前选中的审查列列索引(0-based)
        self.selection_mode = None   # "green" / "red" / None

        # 自动生成的规则列表: [{"green": set, "red": int, "sheets": [name,...]}, ...]
        self.auto_rules = []
        self.rule_widgets = []       # 规则卡片UI控件

        # 列最多的Sheet及其列数
        self.max_sheet = ""
        self.max_cols = 0
        for sn, headers in all_headers_cache.items():
            if len(headers) > self.max_cols:
                self.max_cols = len(headers)
                self.max_sheet = sn

        # 创建弹窗
        self.dialog = ctk.CTkToplevel(parent)
        self.dialog.title("自动选择")
        self.dialog.geometry("750x620")
        self.dialog.minsize(650, 500)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.dialog.configure(fg_color=COLOR_BG)

        self._build_ui()

    def _build_ui(self):
        dialog = self.dialog
        dialog.grid_columnconfigure(0, weight=1)
        dialog.grid_rowconfigure(2, weight=1)  # 下方规则列表区可伸缩

        # ========== 上方：列容器 ==========
        top_frame = ctk.CTkFrame(dialog, fg_color=COLOR_BG_WHITE,
                                 border_width=1, border_color=COLOR_BORDER,
                                 corner_radius=8)
        top_frame.grid(row=0, column=0, sticky="ew", padx=15, pady=(15, 5))

        top_label = ctk.CTkLabel(top_frame, text=f"列预览（来自「{self.max_sheet}」，共{self.max_cols}列）",
                                 font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, family=FONT_FAMILY),
                                 text_color=COLOR_TEXT_MUTED)
        top_label.pack(anchor="w", padx=10, pady=(8, 4))

        # 列单元格滚动容器
        self.col_scroll = ctk.CTkScrollableFrame(top_frame, fg_color="transparent", height=140)
        self.col_scroll.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        self.column_cells = {}  # {col_idx: outer_frame}

        headers = self.all_headers_cache[self.max_sheet]
        # 每行显示10列，超出自动换行，利用垂直滚动全部可见
        cols_per_row = 10
        for idx in range(len(headers)):
            col_letter = get_column_letter(idx + 1)
            cell = ctk.CTkFrame(self.col_scroll, corner_radius=6,
                                cursor="hand2", border_width=0, fg_color=COLOR_BG,
                                width=55, height=42)
            label = ctk.CTkLabel(cell, text=col_letter,
                                font=ctk.CTkFont(size=FONT_SIZE_BODY, weight="bold", family=FONT_FAMILY),
                                text_color=COLOR_TEXT_PRIMARY, width=50, height=38)
            label.pack(padx=4, pady=4)
            cell.bind("<Button-1>", lambda e, i=idx: self._on_column_click(i))
            label.bind("<Button-1>", lambda e, i=idx: self._on_column_click(i))
            # 使用grid网格布局，每cols_per_row列自动换行
            row = idx // cols_per_row
            col = idx % cols_per_row
            cell.grid(row=row, column=col, padx=2, pady=2)
            self.column_cells[idx] = cell

        # ========== 中间：三个按钮 ==========
        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.grid(row=1, column=0, sticky="ew", padx=15, pady=8)
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)
        btn_frame.grid_columnconfigure(2, weight=1)

        self.green_btn = ctk.CTkButton(
            btn_frame, text="选择整体范围",
            command=lambda: self._set_selection_mode("green"),
            font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, family=FONT_FAMILY),
            fg_color="transparent", hover_color=COLOR_GREEN_LIGHT,
            text_color=COLOR_GREEN, border_width=2, border_color=COLOR_GREEN,
            corner_radius=6
        )
        self.green_btn.grid(row=0, column=0, padx=(0, 3), sticky="ew")

        self.red_btn = ctk.CTkButton(
            btn_frame, text="需要审查的列",
            command=lambda: self._set_selection_mode("red"),
            font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, family=FONT_FAMILY),
            fg_color="transparent", hover_color=COLOR_RED_LIGHT,
            text_color=COLOR_RED, border_width=2, border_color=COLOR_RED,
            corner_radius=6
        )
        self.red_btn.grid(row=0, column=1, padx=3, sticky="ew")

        self.auto_btn = ctk.CTkButton(
            btn_frame, text="自动选择",
            command=self._auto_select,
            font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, family=FONT_FAMILY),
            fg_color=COLOR_BLUE, hover_color="#0D47A1",
            text_color="white", corner_radius=6
        )
        self.auto_btn.grid(row=0, column=2, padx=(3, 0), sticky="ew")

        # ========== 下方：已选规则列表 ==========
        bottom_frame = ctk.CTkFrame(dialog, fg_color=COLOR_BG_WHITE,
                                    border_width=1, border_color=COLOR_BORDER,
                                    corner_radius=8)
        bottom_frame.grid(row=2, column=0, sticky="nsew", padx=15, pady=(5, 10))
        bottom_frame.grid_columnconfigure(0, weight=1)
        bottom_frame.grid_rowconfigure(1, weight=1)

        bl_title = ctk.CTkLabel(bottom_frame, text="自动选择的规则",
                                font=ctk.CTkFont(size=FONT_SIZE_TITLE, weight="bold", family=FONT_FAMILY),
                                text_color=COLOR_TEXT_PRIMARY, anchor="w")
        bl_title.grid(row=0, column=0, sticky="w", padx=12, pady=(10, 5))

        self.rules_scroll = ctk.CTkScrollableFrame(bottom_frame, fg_color="transparent")
        self.rules_scroll.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0, 8))

        self.rules_placeholder = ctk.CTkLabel(
            self.rules_scroll, text="点击上方按钮选择列后，按「自动选择」生成规则",
            font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, family=FONT_FAMILY),
            text_color=COLOR_TEXT_MUTED
        )
        self.rules_placeholder.pack(pady=15)

        # ========== 底部按钮栏 ==========
        action_bar = ctk.CTkFrame(dialog, fg_color="transparent")
        action_bar.grid(row=3, column=0, sticky="e", padx=15, pady=(0, 15))
        action_bar.grid_columnconfigure(0, weight=1)

        cancel_btn = ctk.CTkButton(action_bar, text="取消", command=self._on_cancel,
                                   width=80, font=ctk.CTkFont(size=FONT_SIZE_BTN, family=FONT_FAMILY),
                                   fg_color=COLOR_BG, text_color=COLOR_TEXT_PRIMARY,
                                   border_width=1, border_color=COLOR_BORDER,
                                   hover_color=COLOR_RED_LIGHT, corner_radius=6)
        cancel_btn.grid(row=0, column=0, padx=(0, 8), sticky="e")

        ok_btn = ctk.CTkButton(action_bar, text="确定", command=self._on_confirm,
                               width=80, font=ctk.CTkFont(size=FONT_SIZE_BTN, family=FONT_FAMILY),
                               fg_color=COLOR_BLUE, hover_color="#0D47A1",
                               text_color="white", corner_radius=6)
        ok_btn.grid(row=0, column=1, sticky="e")

    def _set_selection_mode(self, mode):
        """设置当前选择模式"""
        if self.selection_mode == mode:
            self.selection_mode = None
        else:
            self.selection_mode = mode
        self._update_mode_buttons()

    def _update_mode_buttons(self):
        """更新模式按钮外观"""
        if self.selection_mode == "green":
            self.green_btn.configure(fg_color=COLOR_GREEN_LIGHT, text_color=COLOR_GREEN)
            self.red_btn.configure(fg_color="transparent", text_color=COLOR_RED)
        elif self.selection_mode == "red":
            self.green_btn.configure(fg_color="transparent", text_color=COLOR_GREEN)
            self.red_btn.configure(fg_color=COLOR_RED_LIGHT, text_color=COLOR_RED)
        else:
            self.green_btn.configure(fg_color="transparent", text_color=COLOR_GREEN)
            self.red_btn.configure(fg_color="transparent", text_color=COLOR_RED)

    def _on_column_click(self, col_idx):
        """点击列单元格"""
        if not self.selection_mode:
            return
        if self.selection_mode == "green":
            if col_idx in self.green_cols:
                self.green_cols.discard(col_idx)
            else:
                self.green_cols.add(col_idx)
        elif self.selection_mode == "red":
            if self.red_col == col_idx:
                self.red_col = None
            else:
                self.red_col = col_idx
        self._redraw_columns()

    def _redraw_columns(self):
        """重绘列单元格边框"""
        for col_idx, cell in self.column_cells.items():
            is_green = col_idx in self.green_cols
            is_red = col_idx == self.red_col
            if is_green:
                cell.configure(border_width=3, border_color=COLOR_GREEN)
            else:
                cell.configure(border_width=0)

            # 红色用内部标记 - 这里简化为直接在边框上叠加显示
            # 用一个简单的方案：如果既是绿色又是红色，红色优先
            if is_red and not is_green:
                cell.configure(border_width=3, border_color=COLOR_RED)
            elif is_red and is_green:
                # 同时选中两种时显示混合（绿底红边）
                cell.configure(border_width=3, border_color=COLOR_RED)

    def _auto_select(self):
        """根据当前选择自动生成所有Sheet的规则"""
        if not self.green_cols:
            messagebox.showwarning("提示", "请先选择整体范围！", parent=self.dialog)
            return
        if self.red_col is None:
            messagebox.showwarning("提示", "请先选择需要审查的列！", parent=self.dialog)
            return

        # 为每个Sheet生成一条规则（对列数不足的Sheet自动截断）
        new_rules = []
        for sheet_name in self.sheet_names:
            headers = self.all_headers_cache[sheet_name]
            max_idx = len(headers) - 1
            # 过滤该sheet能覆盖到的绿色列
            valid_green = {c for c in self.green_cols if c <= max_idx}
            if not valid_green:
                continue
            # 红色列若超出范围则截断为该sheet的最大列索引
            original_red = self.red_col
            adjusted_red = min(self.red_col, max_idx)
            new_rules.append({
                "green_selections": {sheet_name: copy.deepcopy(valid_green)},
                "red_selections": {sheet_name: {adjusted_red}},
                "sheet_name": sheet_name,
                "green_cols": sorted(valid_green),
                "red_col": adjusted_red,
                "original_red": original_red,   # 记录用户原始选择
                "was_adjusted": adjusted_red != original_red  # 是否被截断
            })

        self.auto_rules = new_rules
        self._refresh_rules_list()

        # 汇总提示
        adjusted_count = sum(1 for r in new_rules if r.get("was_adjusted", False))
        if adjusted_count > 0:
            adjusted_names = [r['sheet_name'] for r in new_rules if r.get("was_adjusted")]
            self._show_adjustment_warning(adjusted_count, adjusted_names)
        else:
            messagebox.showinfo("自动选择完成",
                f"已为 {len(new_rules)} 个Sheet生成规则，\n所有Sheet的审查列均未截断，可直接跨Sheet去重。",
                parent=self.dialog)

    def _show_adjustment_warning(self, count, names):
        """显示审查列被截断的警告"""
        detail = "\n".join(names[:10])
        if len(names) > 10:
            detail += f"\n... 等共 {count} 个Sheet"
        messagebox.showwarning("审查列已自动调整",
            f"以下 {count} 个Sheet的列数不足，审查列已自动截断：\n\n{detail}\n\n"
            f"⚠️ 这些Sheet将读取不同的物理列作为审查列，\n可能导致跨Sheet去重不彻底！\n"
            f"建议：检查这些Sheet的数据布局是否一致。", parent=self.dialog)

    def _refresh_rules_list(self):
        """刷新下方规则列表UI"""
        for w in self.rule_widgets:
            w.destroy()
        self.rule_widgets.clear()

        if not self.auto_rules:
            self.rules_placeholder.pack(pady=15)
            return

        self.rules_placeholder.pack_forget()

        for i, rule in enumerate(self.auto_rules):
            card = ctk.CTkFrame(self.rules_scroll, fg_color=COLOR_BG_WHITE,
                               border_width=1, border_color=COLOR_BORDER, corner_radius=8)
            card.pack(fill="x", pady=4, padx=4)

            inner = ctk.CTkFrame(card, fg_color="transparent")
            inner.pack(fill="both", expand=True, padx=8, pady=6)
            inner.grid_columnconfigure(0, weight=1)

            # 标题行：Sheet名 + 删除按钮
            title_row = ctk.CTkFrame(inner, fg_color="transparent")
            title_row.grid(row=0, column=0, sticky="ew", pady=(0, 4))
            title_row.grid_columnconfigure(0, weight=1)

            title_lbl = ctk.CTkLabel(title_row, text=f"规则 {i+1}：{rule['sheet_name']}",
                                     font=ctk.CTkFont(size=FONT_SIZE_BODY, weight="bold", family=FONT_FAMILY),
                                     text_color=COLOR_TEXT_PRIMARY, anchor="w")
            title_lbl.grid(row=0, column=0, sticky="w")

            del_btn = ctk.CTkButton(title_row, text="×删除", width=60, height=26,
                                   font=ctk.CTkFont(size=FONT_SIZE_SMALL, family=FONT_FAMILY),
                                   fg_color=COLOR_RED_LIGHT, hover_color="#EF5350",
                                   text_color=COLOR_RED, corner_radius=5,
                                   command=lambda idx=i: self._delete_rule(idx))
            del_btn.grid(row=0, column=1, sticky="e")

            # 内容行：绿色范围 + 红色列
            content_row = ctk.CTkFrame(inner, fg_color="transparent")
            content_row.grid(row=1, column=0, sticky="ew")

            green_letters = [get_column_letter(c + 1) for c in rule["green_cols"]]
            red_letter = get_column_letter(rule["red_col"] + 1)
            was_adjusted = rule.get("was_adjusted", False)

            green_area = ctk.CTkFrame(content_row, fg_color=COLOR_GREEN_LIGHT, corner_radius=5)
            green_area.pack(side="left", fill="both", expand=True, padx=(0, 3))
            green_label = ctk.CTkLabel(green_area,
                                       text=f"🟢 整体范围\n{', '.join(green_letters)}",
                                       font=ctk.CTkFont(size=FONT_SIZE_SMALL, family=FONT_FAMILY),
                                       text_color=COLOR_TEXT_PRIMARY, anchor="center")
            green_label.pack(padx=6, pady=4)

            # 审查列：若被截断则显示警告
            if was_adjusted:
                orig_letter = get_column_letter(rule["original_red"] + 1)
                red_text = f"⚠️ 审查列\n{red_letter} (原选{orig_letter})"
                red_text_color = COLOR_ORANGE
            else:
                red_text = f"🔴 审查列\n{red_letter}"
                red_text_color = COLOR_TEXT_PRIMARY

            red_area = ctk.CTkFrame(content_row, fg_color=COLOR_RED_LIGHT, corner_radius=5)
            red_area.pack(side="left", fill="both", expand=True, padx=(3, 0))
            red_label = ctk.CTkLabel(red_area,
                                     text=red_text,
                                     font=ctk.CTkFont(size=FONT_SIZE_SMALL, family=FONT_FAMILY),
                                     text_color=red_text_color, anchor="center")
            red_label.pack(padx=6, pady=4)

            self.rule_widgets.append(card)

    def _delete_rule(self, rule_idx):
        """删除指定规则"""
        if 0 <= rule_idx < len(self.auto_rules):
            self.auto_rules.pop(rule_idx)
            self._refresh_rules_list()

    def _on_cancel(self):
        """取消按钮"""
        self.dialog.destroy()

    def _on_confirm(self):
        """确定按钮 -> 返回结果并关闭"""
        if not self.auto_rules:
            messagebox.showwarning("提示", "请先生成至少一条规则！", parent=self.dialog)
            return
        self.result = self.auto_rules
        self.dialog.destroy()


class ExcelDeduplicationTool:
    """Excel跨Sheet单元格去重工具主类"""

    # ==================== 统一异常处理 ====================

    @staticmethod
    def _classify_error(e, context="", file_path=None):
        """
        将原始Python异常转换为用户友好的业务异常（ExcelCleanerError）
        返回 (business_error, log_detail) 元组
        """
        # ---- 已是自定义异常 → 直接返回 ----
        if isinstance(e, ExcelCleanerError):
            return e, str(e)

        e_type = type(e).__name__
        e_msg = str(e).strip()
        detail = f"[{e_type}] {e_msg}" if e_msg else f"[{e_type}]"

        # ---- 文件不存在 ----
        if isinstance(e, FileNotFoundError) or "No such file" in e_msg:
            return FileNotFoundError_(file_path or "未知路径"), detail

        # ---- 权限问题 / 文件被占用 ----
        if isinstance(e, PermissionError):
            is_locked = any(kw in e_msg.lower() for kw in ["used by another process", "being used", "access is denied", "权限", "被占用"])
            return FilePermissionError_(file_path or "未知路径", is_locked=is_locked), detail

        # ---- Excel文件损坏 / 格式错误 ----
        if isinstance(e, InvalidFileException) or isinstance(e, BadZipFile):
            return FileCorruptError(file_path or "", detail=e_msg), detail

        if "not a valid zip file" in e_msg or "is not a zip file" in e_msg:
            return FileCorruptError(file_path or "", detail="文件不是有效的zip格式（xlsx底层是zip）"), detail

        if "cannot open" in e_msg.lower() or "invalid signature" in e_msg.lower():
            return FileCorruptError(file_path or "", detail=e_msg), detail

        # ---- 内存不足 ----
        if isinstance(e, MemoryError):
            return MemoryError_(), detail

        # ---- 磁盘空间不足 ----
        if isinstance(e, OSError) and ("no space left" in e_msg.lower() or "磁盘空间不足" in e_msg):
            return DiskSpaceError(required_mb=0, available_mb=0), detail

        # ---- Sheet不存在（KeyError在openpyxl访问sheet时触发）----
        if isinstance(e, KeyError):
            sheet_hint = f"（可能是工作表「{e.args[0]}」不存在）" if e.args else ""
            return SheetNotFoundError(str(e.args[0]) if e else "未知"), detail + sheet_hint

        # ---- 索引越界 ----
        if isinstance(e, IndexError):
            return ColumnOutOfRangeError(col_idx=-1, max_col=-1, detail=e_msg), detail

        # ---- 写入/保存失败 ----
        if isinstance(e, IOError) or isinstance(e, OSError):
            if any(kw in e_msg.lower() for kw in ["write", "save", "保存", "写入"]):
                return WriteFileError(file_path or "", detail=e_msg), detail

        # ---- 默认：包装为通用异常 ----
        generic = ExcelCleanerError(
            message=f"{context}失败：{e_msg}",
            error_code="ERR",
            suggestion=f"请检查操作是否正确，如问题持续存在请联系技术支持。\n技术详情：{detail}"
        )
        return generic, detail

    def _safe_log_error(self, error_obj, detail=""):
        """安全地记录错误日志到UI"""
        code = getattr(error_obj, 'error_code', '?')
        msg = getattr(error_obj, 'message', str(error_obj))
        suggestion = getattr(error_obj, 'suggestion', None)

        self._log(f"\n{'='*60}")
        self._log(f"【错误 {code}】")
        self._log(f"  描述：{msg}")
        if suggestion:
            for line in suggestion.split('\n'):
                self._log(f"  建议：{line}")
        if detail:
            self._log(f"  技术详情：{detail}")
        self._log(f"{'='*60}")

    def __init__(self, root):
        """初始化主窗口"""
        self.root = root
        self.root.title("Excel数据清洗工具")
        self.root.geometry("1100x680")
        self.root.minsize(900, 550)

        # 设置窗口背景色
        self.root.configure(fg_color=COLOR_BG)

        # 数据变量
        self.file_path = ctk.StringVar()
        self.skip_header_var = ctk.BooleanVar(value=True)
        self.rule_mode_var = ctk.StringVar(value="single")  # "single" 或 "multi"

        # 处理控制
        self._pause_event = threading.Event()   # set=运行中, clear=暂停中
        self._cancel_flag = False               # True=取消处理

        # Sheet 相关
        self.sheet_names = []            # 当前文件的Sheet名称列表
        self.selected_sheet = ctk.StringVar(value="")

        # 列选择状态
        self.green_selections = {}       # 跨Sheet整体范围：{sheet_name: set(col_indices)}
        self.red_selections = {}         # 跨Sheet审查列：{sheet_name: set(col_indices)}
        self.selection_mode = None       # 当前选择模式："green" / "red" / None
        self.column_headers = []         # 当前显示Sheet的表头列表
        self.column_cells = {}           # {col_idx: (outer_frame, inner_frame)}
        self.column_cells_frame = None   # 列单元格容器
        self._col_configure_id = None    # Configure事件绑定ID
        self._relayout_after_id = None   # 防抖回调ID

        # 已保存的规则列表：[{green_selections: {sheet: set}, red_selections: {sheet: set}}, ...]
        self.saved_rules = []
        self.rule_widgets = []           # 每条规则对应的UI框架列表

        # 计时器相关
        self.timer_running = False
        self.process_start_time = 0
        self.timer_after_id = None

        # 日志缓冲区（用于处理完成后导出完整日志文件）
        self._log_buffer = []           # Python _log() 的所有输出
        self._go_log_content = ""       # Go 程序的 stderr 完整内容
        self._processing_file_path = "" # 当前正在处理的源文件路径
        self._is_closing = False        # 窗口是否正在关闭
        self._trace_keywords = []       # Go引擎关键词追踪列表（用于诊断）

        # 构建界面
        self._build_ui()

        # 注册窗口关闭回调，防止后台线程/定时器在窗口销毁后调用tkinter
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)

    def _build_ui(self):
        """构建可视化界面（左右布局）"""
        self.root.grid_columnconfigure(0, weight=3)  # 左侧配置区
        self.root.grid_columnconfigure(1, weight=2)  # 右侧日志区
        self.root.grid_rowconfigure(0, weight=1)     # 主体区域可伸缩

        # ==================== 左侧面板：文件选择 + 配置 ====================
        left_panel = ctk.CTkFrame(
            self.root, fg_color=COLOR_BG_WHITE,
            border_width=2, border_color="#C8CED3",
            corner_radius=10
        )
        left_panel.grid(row=0, column=0, padx=(15, 5), pady=15, sticky="nsew")
        left_panel.grid_columnconfigure(0, weight=1)
        left_panel.grid_rowconfigure(1, weight=1)

        # ---- 1. 文件选择区域 ----
        file_frame = ctk.CTkFrame(left_panel, fg_color="transparent")
        file_frame.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 5))
        file_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(file_frame, text="文件选择",
                     font=ctk.CTkFont(size=FONT_SIZE_TITLE, weight="bold", family=FONT_FAMILY),
                     text_color=COLOR_TEXT_PRIMARY).grid(row=0, column=0, sticky="w", pady=(0, 3))

        ctk.CTkLabel(file_frame, text="当前版本：V1.0",
                     font=ctk.CTkFont(size=12, family=FONT_FAMILY),
                     text_color=COLOR_TEXT_MUTED).grid(row=0, column=1, sticky="e", pady=(0, 3))

        entry_row = ctk.CTkFrame(file_frame, fg_color="transparent")
        entry_row.grid(row=1, column=0, sticky="ew")
        entry_row.grid_columnconfigure(0, weight=1)

        file_entry = ctk.CTkEntry(entry_row, textvariable=self.file_path,
                                  placeholder_text="请选择Excel文件...",
                                  font=ctk.CTkFont(size=FONT_SIZE_BODY, family=FONT_FAMILY),
                                  fg_color=COLOR_BG, border_width=1, border_color=COLOR_BORDER,
                                  text_color=COLOR_TEXT_PRIMARY, placeholder_text_color=COLOR_TEXT_MUTED,
                                  corner_radius=6)
        file_entry.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        self._file_entry = file_entry  # 保存引用，处理时禁用

        select_btn = ctk.CTkButton(entry_row, text="选择文件", command=self._select_file,
                                   width=90, font=ctk.CTkFont(size=FONT_SIZE_BTN, family=FONT_FAMILY),
                                   fg_color=COLOR_BLUE, hover_color="#0D47A1",
                                   text_color="white", corner_radius=6)
        select_btn.grid(row=0, column=1)
        self._select_btn = select_btn  # 保存引用，处理时禁用

        # ---- 2. 配置区域 ----
        config_frame = ctk.CTkScrollableFrame(
            left_panel, label_text="配置区域",
            fg_color=COLOR_BG, border_width=1, border_color=COLOR_BORDER,
            corner_radius=8,
            label_font=ctk.CTkFont(size=FONT_SIZE_TITLE, weight="bold", family=FONT_FAMILY),
            label_text_color=COLOR_TEXT_PRIMARY
        )
        config_frame.grid(row=1, column=0, sticky="nsew", pady=(0, 5), padx=12)
        config_frame.grid_columnconfigure(0, weight=1)
        # 配置各行的权重,让规则列表区域(row=6)占据剩余空间
        for i in range(7):
            config_frame.grid_rowconfigure(i, weight=0)
        config_frame.grid_rowconfigure(6, weight=1)

        # 标题行 + 跳过表头开关
        title_row = ctk.CTkFrame(config_frame, fg_color="transparent")
        title_row.grid(row=0, column=0, sticky="ew", pady=(5, 5))
        title_row.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(title_row, text="跳过表头",
                     font=ctk.CTkFont(size=FONT_SIZE_BODY, family=FONT_FAMILY),
                     text_color=COLOR_TEXT_PRIMARY).grid(row=0, column=0, sticky="w")

        self.skip_header_switch = ctk.CTkSwitch(
            title_row, text="", variable=self.skip_header_var,
            onvalue=True, offvalue=False,
            fg_color=COLOR_BORDER, progress_color=COLOR_BLUE,
            button_color=COLOR_BG_WHITE, button_hover_color=COLOR_BG
        )
        self.skip_header_switch.grid(row=0, column=1, sticky="e")

        # 规则审查模式选择行
        rule_mode_row = ctk.CTkFrame(config_frame, fg_color="transparent")
        rule_mode_row.grid(row=1, column=0, sticky="ew", pady=(0, 5))
        rule_mode_row.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(rule_mode_row, text="审查模式：",
                     font=ctk.CTkFont(size=FONT_SIZE_BODY, family=FONT_FAMILY),
                     text_color=COLOR_TEXT_PRIMARY).grid(row=0, column=0, sticky="w")

        # 分段按钮
        self.rule_mode_segbtn = ctk.CTkSegmentedButton(
            rule_mode_row,
            values=["单规则审查", "多规则联查"],
            font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, family=FONT_FAMILY),
            command=self._on_rule_mode_changed,
            fg_color=COLOR_BG_WHITE,
            selected_color=COLOR_BLUE, selected_hover_color="#0D47A1",
            text_color=COLOR_TEXT_PRIMARY
        )
        self.rule_mode_segbtn.grid(row=0, column=1, sticky="ew", padx=(8, 5))
        # 设置初始选中
        self.rule_mode_segbtn.set("单规则审查")

        # 帮助图标按钮(点击弹出说明)
        self.rule_mode_help_btn = ctk.CTkButton(
            rule_mode_row, text="!", width=28, height=28,
            font=ctk.CTkFont(size=14, weight="bold"),
            corner_radius=14,
            fg_color=COLOR_ORANGE, hover_color="#F57C00",
            text_color="white",
            command=self._show_rule_mode_help
        )
        self.rule_mode_help_btn.grid(row=0, column=2, padx=(0, 0))

        # Sheet 选择行（默认隐藏，选择文件后显示）
        self.sheet_row = ctk.CTkFrame(config_frame, fg_color="transparent")
        self.sheet_row.grid(row=2, column=0, sticky="ew", pady=(0, 0))
        self.sheet_row.grid_columnconfigure(1, weight=1)
        self.sheet_row.grid_remove()

        ctk.CTkLabel(self.sheet_row, text="Sheet页：",
                     font=ctk.CTkFont(size=FONT_SIZE_BODY, family=FONT_FAMILY),
                     text_color=COLOR_TEXT_PRIMARY).grid(row=0, column=0, sticky="w")

        self.sheet_optionmenu = ctk.CTkOptionMenu(
            self.sheet_row, variable=self.selected_sheet,
            font=ctk.CTkFont(size=FONT_SIZE_BODY, family=FONT_FAMILY),
            fg_color=COLOR_BG_WHITE, button_color=COLOR_BG,
            button_hover_color=COLOR_BORDER, text_color=COLOR_TEXT_PRIMARY,
            corner_radius=6,
            command=self._on_sheet_changed
        )
        self.sheet_optionmenu.grid(row=0, column=1, sticky="ew", padx=(8, 0))

        # 表头预览区域（占位 或 列单元格）
        self.col_area = ctk.CTkFrame(config_frame, fg_color=COLOR_BG_WHITE,
                                     border_width=1, border_color=COLOR_BORDER,
                                     corner_radius=8)
        self.col_area.grid(row=3, column=0, sticky="ew", pady=(0, 5))

        self.col_placeholder = ctk.CTkLabel(
            self.col_area, text="请先选择Excel文件",
            font=ctk.CTkFont(size=FONT_SIZE_BODY, family=FONT_FAMILY),
            text_color=COLOR_TEXT_MUTED
        )
        self.col_placeholder.pack(padx=10, pady=12)

        # 模式按钮行（2个按钮）
        mode_row = ctk.CTkFrame(config_frame, fg_color="transparent")
        mode_row.grid(row=4, column=0, sticky="ew", pady=(2, 5))
        mode_row.grid_columnconfigure(0, weight=1)
        mode_row.grid_columnconfigure(1, weight=1)

        self.green_mode_btn = ctk.CTkButton(
            mode_row, text="选择整体范围",
            command=lambda: self._set_selection_mode("green"),
            font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, family=FONT_FAMILY),
            fg_color="transparent", hover_color=COLOR_GREEN_LIGHT,
            text_color=COLOR_GREEN,
            border_width=2, border_color=COLOR_GREEN,
            corner_radius=6
        )
        self.green_mode_btn.grid(row=0, column=0, padx=(0, 3), sticky="ew")

        self.red_mode_btn = ctk.CTkButton(
            mode_row, text="需要审查的列",
            command=lambda: self._set_selection_mode("red"),
            font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, family=FONT_FAMILY),
            fg_color="transparent", hover_color=COLOR_RED_LIGHT,
            text_color=COLOR_RED,
            border_width=2, border_color=COLOR_RED,
            corner_radius=6
        )
        self.red_mode_btn.grid(row=0, column=1, padx=(3, 0), sticky="ew")

        # 选择信息行 + 保存按钮 + 自动选择按钮
        info_row = ctk.CTkFrame(config_frame, fg_color="transparent")
        info_row.grid(row=5, column=0, sticky="ew", pady=(0, 5))
        info_row.grid_columnconfigure(0, weight=1)

        self.selection_info = ctk.CTkLabel(
            info_row, text="",
            font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, family=FONT_FAMILY),
            text_color=COLOR_TEXT_MUTED, anchor="w"
        )
        self.selection_info.grid(row=0, column=0, sticky="w")

        self.save_rule_btn = ctk.CTkButton(
            info_row, text="保存规则", command=self._save_rule,
            width=90, font=ctk.CTkFont(size=FONT_SIZE_BTN, family=FONT_FAMILY),
            fg_color=COLOR_BLUE, hover_color="#0D47A1",
            text_color="white", corner_radius=6
        )
        self.save_rule_btn.grid(row=0, column=1, padx=(10, 5))

        self.auto_select_btn = ctk.CTkButton(
            info_row, text="自动选择", command=self._open_auto_select_dialog,
            width=90, font=ctk.CTkFont(size=FONT_SIZE_BTN, family=FONT_FAMILY),
            fg_color=COLOR_ORANGE, hover_color="#F57C00",
            text_color="white", corner_radius=6
        )
        self.auto_select_btn.grid(row=0, column=2)

        # 规则列表区域
        rules_container = ctk.CTkFrame(config_frame, fg_color="transparent")
        rules_container.grid(row=6, column=0, sticky="nsew", pady=(0, 5))
        rules_container.grid_columnconfigure(0, weight=1)
        rules_container.grid_rowconfigure(1, weight=1)
        
        # 规则列表标题行
        rules_header = ctk.CTkFrame(rules_container, fg_color="transparent")
        rules_header.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        rules_header.grid_columnconfigure(0, weight=1)
        
        rules_title = ctk.CTkLabel(
            rules_header, text="已保存的规则",
            font=ctk.CTkFont(size=FONT_SIZE_TITLE, weight="bold", family=FONT_FAMILY),
            text_color=COLOR_TEXT_PRIMARY, anchor="w"
        )
        rules_title.grid(row=0, column=0, sticky="w")
        
        self.clear_rules_btn = ctk.CTkButton(
            rules_header, text="清空列表", width=80,
            font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, family=FONT_FAMILY),
            fg_color=COLOR_BG, text_color=COLOR_TEXT_MUTED,
            border_width=1, border_color=COLOR_BORDER,
            hover_color=COLOR_RED_LIGHT,
            corner_radius=6,
            command=self._clear_all_rules
        )
        self.clear_rules_btn.grid(row=0, column=1, padx=(10, 0))
        
        # 规则列表滚动区域
        self.rules_list_frame = ctk.CTkScrollableFrame(rules_container)
        self.rules_list_frame.grid(row=1, column=0, sticky="nsew")

        # 规则为空时的占位提示
        self.rules_placeholder = ctk.CTkLabel(
            self.rules_list_frame, text="暂无规则，请先制定规则",
            font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, family=FONT_FAMILY),
            text_color=COLOR_TEXT_MUTED
        )
        self.rules_placeholder.pack(pady=8)

        # ---- 3. 底部按钮栏（左侧底部固定） ----
        btn_bar = ctk.CTkFrame(left_panel, fg_color="transparent")
        btn_bar.grid(row=3, column=0, sticky="ew", pady=(5, 12), padx=12)
        btn_bar.grid_columnconfigure(0, weight=1)
        btn_bar.grid_columnconfigure(1, weight=1)
        btn_bar.grid_columnconfigure(2, weight=1)

        self.process_btn = ctk.CTkButton(
            btn_bar, text="开始处理", command=self._start_processing,
            font=ctk.CTkFont(size=FONT_SIZE_BTN + 2, weight="bold", family=FONT_FAMILY),
            height=40, corner_radius=8,
            fg_color=COLOR_BLUE, hover_color="#0D47A1", text_color="white"
        )
        self.process_btn.grid(row=0, column=0, sticky="ew", padx=(0, 3))

        self.pause_btn = ctk.CTkButton(
            btn_bar, text="暂停", command=self._toggle_pause,
            font=ctk.CTkFont(size=FONT_SIZE_BTN + 1, family=FONT_FAMILY),
            height=40, corner_radius=8,
            fg_color=COLOR_YELLOW, hover_color="#F9A825", text_color="white",
            state="disabled"
        )
        self.pause_btn.grid(row=0, column=1, sticky="ew", padx=3)

        self.cancel_btn = ctk.CTkButton(
            btn_bar, text="取消", command=self._cancel_processing,
            font=ctk.CTkFont(size=FONT_SIZE_BTN + 1, family=FONT_FAMILY),
            height=40, corner_radius=8,
            fg_color=COLOR_RED, hover_color="#B71C1C", text_color="white",
            state="disabled"
        )
        self.cancel_btn.grid(row=0, column=2, sticky="ew", padx=(3, 0))

        # ==================== 右侧面板：运行日志 ====================
        log_frame = ctk.CTkFrame(
            self.root, fg_color=COLOR_BG_WHITE,
            border_width=2, border_color="#C8CED3",
            corner_radius=10
        )
        log_frame.grid(row=0, column=1, padx=(5, 15), pady=15, sticky="nsew")

        # 日志标题行（左标题 + 右计时器）
        log_header = ctk.CTkFrame(log_frame, fg_color="transparent")
        log_header.pack(fill="x", padx=12, pady=(12, 5))
        log_header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(log_header, text="处理日志",
                     font=ctk.CTkFont(size=FONT_SIZE_TITLE, weight="bold", family=FONT_FAMILY),
                     text_color=COLOR_TEXT_PRIMARY).grid(row=0, column=0, sticky="w")

        self.timer_label = ctk.CTkLabel(
            log_header, text="已处理：00:00:00 | 剩余：--:--:--",
            font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, family=FONT_MONO),
            text_color=COLOR_TEXT_MUTED
        )
        self.timer_label.grid(row=0, column=1, sticky="e")

        self.log_text = ctk.CTkTextbox(
            log_frame,
            font=ctk.CTkFont(family=FONT_MONO, size=FONT_SIZE_SMALL + 1),
            fg_color=COLOR_BG, text_color=COLOR_TEXT_PRIMARY,
            border_width=1, border_color=COLOR_BORDER,
            corner_radius=8,
            wrap="word", state="disabled"
        )
        self.log_text.pack(fill="both", expand=True, padx=12, pady=(0, 12))

    # ==================== 规则模式相关方法 ====================

    def _on_rule_mode_changed(self, value):
        """规则审查模式切换回调"""
        if value == "单规则审查":
            self.rule_mode_var.set("single")
        else:
            self.rule_mode_var.set("multi")
        self._log(f"已切换审查模式：{value}")

    def _show_rule_mode_help(self):
        """显示规则模式帮助信息"""
        help_msg = (
            "【单规则审查】\n"
            "单Sheet页对指定范围进行审查去重。\n"
            "每条规则独立去重，规则之间互不影响。\n\n"
            "【多规则联查】\n"
            "对多个规则内的所有整体范围进行统一去重。\n"
            "所有规则共享同一个去重记录，全局去重。"
        )
        messagebox.showinfo("审查模式说明", help_msg)

    # ==================== 列选择相关方法 ====================

    def _set_selection_mode(self, mode):
        """切换选择模式"""
        if self.selection_mode == mode:
            self.selection_mode = None
        else:
            self.selection_mode = mode
        self._update_mode_buttons()

    def _update_mode_buttons(self):
        """更新模式按钮的视觉状态"""
        if self.selection_mode == "green":
            # 绿色按钮激活状态:实心深绿背景+白色文字
            self.green_mode_btn.configure(
                fg_color="#1B5E20", hover_color="#0D3D0D",
                border_width=2, border_color="#0D3D0D",
                text_color="white"
            )
            # 红色按钮恢复默认(描边风格)
            self.red_mode_btn.configure(
                fg_color="transparent", hover_color=COLOR_RED_LIGHT,
                border_width=2, border_color=COLOR_RED,
                text_color=COLOR_RED
            )
        elif self.selection_mode == "red":
            # 红色按钮激活状态:实心深红背景+白色文字
            self.red_mode_btn.configure(
                fg_color="#B71C1C", hover_color="#7F0000",
                border_width=2, border_color="#7F0000",
                text_color="white"
            )
            # 绿色按钮恢复默认(描边风格)
            self.green_mode_btn.configure(
                fg_color="transparent", hover_color=COLOR_GREEN_LIGHT,
                border_width=2, border_color=COLOR_GREEN,
                text_color=COLOR_GREEN
            )
        else:
            # 两个按钮都恢复默认(描边风格)
            self.green_mode_btn.configure(
                fg_color="transparent", hover_color=COLOR_GREEN_LIGHT,
                border_width=2, border_color=COLOR_GREEN,
                text_color=COLOR_GREEN
            )
            self.red_mode_btn.configure(
                fg_color="transparent", hover_color=COLOR_RED_LIGHT,
                border_width=2, border_color=COLOR_RED,
                text_color=COLOR_RED
            )
        self._update_selection_info()

    def _on_column_click(self, col_idx):
        """点击列单元格的回调（仅当前Sheet）"""
        if not self.selection_mode:
            return

        current_sheet = self.selected_sheet.get()
        if not current_sheet:
            return

        # 清除其他Sheet的选择,只保留当前Sheet
        if self.selection_mode == "green":
            # 只保留当前sheet的选择
            other_selections = {k: v for k, v in self.green_selections.items() if k == current_sheet}
            self.green_selections.clear()
            self.green_selections.update(other_selections)
            
            if current_sheet not in self.green_selections:
                self.green_selections[current_sheet] = set()
            cols = self.green_selections[current_sheet]
            if col_idx in cols:
                cols.discard(col_idx)
                if not cols:
                    del self.green_selections[current_sheet]
            else:
                cols.add(col_idx)
        elif self.selection_mode == "red":
            # 红色模式:只允许选择单列
            other_selections = {k: v for k, v in self.red_selections.items() if k == current_sheet}
            self.red_selections.clear()
            self.red_selections.update(other_selections)
            
            if current_sheet not in self.red_selections:
                self.red_selections[current_sheet] = set()
            cols = self.red_selections[current_sheet]
            
            # 如果点击的是已选择的列,则取消选择
            if col_idx in cols:
                cols.discard(col_idx)
                if not cols:
                    del self.red_selections[current_sheet]
            else:
                # 清除之前的选择,只保留新选择的列(单选)
                cols.clear()
                cols.add(col_idx)

        self._redraw_column_cells()
        self._update_selection_info()

    def _redraw_column_cells(self):
        """根据当前Sheet的选择状态重绘列单元格边框"""
        current_sheet = self.selected_sheet.get()
        green_cols = self.green_selections.get(current_sheet, set())
        red_cols = self.red_selections.get(current_sheet, set())

        for col_idx, (outer, inner) in self.column_cells.items():
            is_green = col_idx in green_cols
            is_red = col_idx in red_cols

            if is_green:
                outer.configure(border_width=3, border_color=COLOR_GREEN)
            else:
                outer.configure(border_width=0)

            if is_red:
                inner.configure(border_width=3, border_color=COLOR_RED)
            else:
                inner.configure(border_width=0)

    def _update_selection_info(self):
        """更新底部选择信息文字(仅显示当前Sheet)"""
        current_sheet = self.selected_sheet.get()
        if not current_sheet:
            self.selection_info.configure(text="")
            return
        
        parts = []
        if current_sheet in self.green_selections and self.green_selections[current_sheet]:
            cols = sorted(self.green_selections[current_sheet])
            letters = [get_column_letter(c + 1) for c in cols]
            parts.append(f"🟢 整体范围：{', '.join(letters)}")
        if current_sheet in self.red_selections and self.red_selections[current_sheet]:
            cols = sorted(self.red_selections[current_sheet])
            letters = [get_column_letter(c + 1) for c in cols]
            parts.append(f"🔴 审查的列：{', '.join(letters)}")
        self.selection_info.configure(text="   ".join(parts) if parts else "")

    def _on_sheet_changed(self, value):
        """切换Sheet时保留各Sheet的选择状态，仅刷新列显示"""
        # 不再清空选择状态！每个Sheet独立保存自己的绿色/红色选择
        self.selection_mode = None  # 只重置当前模式按钮
        self._update_mode_buttons()
        self._smart_update_column_display()

    def _read_headers(self, file_path):
        """读取选中Sheet的第一行，返回表头列表（复用缓存的workbook）"""
        try:
            if not hasattr(self, '_ro_wb') or self._ro_wb is None:
                self._ro_wb = load_workbook(filename=file_path, read_only=True)
                self._ro_wb_path = file_path
            elif self._ro_wb_path != file_path:
                try:
                    self._ro_wb.close()
                except Exception:
                    pass
                self._ro_wb = load_workbook(filename=file_path, read_only=True)
                self._ro_wb_path = file_path
            sheet_name = self.selected_sheet.get()
            ws = self._ro_wb[sheet_name] if sheet_name in self._ro_wb.sheetnames else self._ro_wb.active
            headers = []
            for row in ws.iter_rows(max_row=1, values_only=True):
                if row:
                    headers = list(row)
                break
            return headers
        except InvalidFileException as e:
            # E003: 文件损坏或格式不兼容
            biz_err, detail = self._classify_error(e, context="读取Excel", file_path=file_path)
            self._safe_log_error(biz_err, detail)
            return None
        except PermissionError as e:
            # E002: 文件被占用或无权限
            biz_err, detail = self._classify_error(e, context="读取文件", file_path=file_path)
            self._safe_log_error(biz_err, detail)
            messagebox.showerror(f"错误 {biz_err.error_code}", biz_err.get_user_message())
            return None
        except FileNotFoundError as e:
            # E001: 文件不存在
            biz_err, detail = self._classify_error(e, context="打开文件", file_path=file_path)
            self._safe_log_error(biz_err, detail)
            return None
        except BadZipFile as e:
            # E003: 非标准xlsx文件
            biz_err, detail = self._classify_error(e, context="解析Excel", file_path=file_path)
            self._safe_log_error(biz_err, detail)
            return None
        except Exception as e:
            # 其他未知错误，记录日志但不中断UI
            self._log(f"【警告】读取表头时发生意外错误：{type(e).__name__} - {str(e)}")
            return None

    def _close_ro_wb(self):
        """关闭缓存的只读workbook"""
        if hasattr(self, '_ro_wb') and self._ro_wb is not None:
            try:
                self._ro_wb.close()
            except Exception:
                pass
            self._ro_wb = None
            self._ro_wb_path = None

    def _init_column_display_with_headers(self, headers):
        """使用已有的表头数据初始化列显示（跳过文件读取）"""
        # 取消旧的Configure监听
        if hasattr(self, '_col_configure_id') and self._col_configure_id:
            self.col_area.unbind("<Configure>")
            self._col_configure_id = None

        # 清除旧的列单元格
        if self.column_cells_frame is not None:
            self.column_cells_frame.destroy()
            self.column_cells_frame = None
        self.column_cells.clear()

        if not headers:
            self.col_placeholder.configure(text="无法读取表头，请确认文件格式")
            self.col_placeholder.pack(padx=10, pady=12)
            return

        self.column_headers = headers
        self.col_placeholder.pack_forget()

        # 创建列单元格容器
        self.column_cells_frame = ctk.CTkFrame(self.col_area, fg_color="transparent")
        self.column_cells_frame.pack(fill="both", expand=True, padx=5, pady=8)

        # 先创建所有单元格（暂不布局）
        for idx, header in enumerate(headers):
            display = str(header).strip() if header else ""
            if not display:
                display = f"(空)"
            if len(display) > 6:
                display = display[:5] + "…"
            col_letter = get_column_letter(idx + 1)
            text = f"{col_letter}\n{display}"

            outer = ctk.CTkFrame(self.column_cells_frame, corner_radius=6,
                                 cursor="hand2", border_width=0,
                                 fg_color=COLOR_BG)
            inner = ctk.CTkFrame(outer, corner_radius=4, border_width=0,
                                 fg_color=COLOR_BG_WHITE, border_color=COLOR_BORDER)
            label = ctk.CTkLabel(inner, text=text,
                                 font=ctk.CTkFont(size=FONT_SIZE_SMALL, family=FONT_FAMILY),
                                 text_color=COLOR_TEXT_PRIMARY,
                                 width=75, height=42, anchor="center")
            label.pack(padx=4, pady=4)
            inner.pack(padx=3, pady=3)

            for widget in (outer, inner, label):
                widget.bind("<Button-1>", lambda e, i=idx: self._on_column_click(i))

            self.column_cells[idx] = (outer, inner)

        # 初始布局 + 监听宽度变化自动重排
        self.root.after_idle(self._relayout_columns)
        self._col_configure_id = self.col_area.bind(
            "<Configure>", lambda e: self.root.after_idle(self._relayout_columns)
        )

        # 应用已有选择状态
        self._redraw_column_cells()
        self._update_selection_info()

    def _cleanup_file_state(self):
        """切换文件时清理上一份文件的所有状态（内存/缓存）"""
        # 1. 停止正在运行的计时器
        if self.timer_after_id:
            try:
                self.root.after_cancel(self.timer_after_id)
            except Exception:
                pass
            self.timer_after_id = None
        self.timer_running = False
        self.process_start_time = 0

        # 2. 取消正在进行的处理
        if self._cancel_flag is False and not self._pause_event.is_set():
            pass  # 没有正在运行的处理
        self._cancel_flag = False
        self._pause_event.set()

        # 3. 防抖回调
        if self._relayout_after_id:
            try:
                self.root.after_cancel(self._relayout_after_id)
            except Exception:
                pass
            self._relayout_after_id = None

        # 4. 清理列显示相关
        if hasattr(self, '_col_configure_id') and self._col_configure_id:
            try:
                self.col_area.unbind("<Configure>")
            except Exception:
                pass
            self._col_configure_id = None
        if self.column_cells_frame is not None:
            self.column_cells_frame.destroy()
            self.column_cells_frame = None
        self.column_cells.clear()
        self.column_headers.clear()

        # 5. 清理列选择状态
        self.green_selections.clear()
        self.red_selections.clear()
        self.selection_mode = None

        # 6. 清理已保存的规则
        self.saved_rules.clear()
        for w in self.rule_widgets:
            try:
                w.destroy()
            except Exception:
                pass
        self.rule_widgets.clear()
        if hasattr(self, 'rules_placeholder') and self.rules_placeholder:
            try:
                self.rules_placeholder.pack(pady=8)
            except Exception:
                pass

        # 7. 重置选择信息
        if hasattr(self, 'selection_info'):
            self.selection_info.configure(text="")

        # 8. 关闭缓存的只读workbook
        self._close_ro_wb()

        # 8.5 清理表头缓存
        if hasattr(self, '_all_sheet_headers_cache'):
            del self._all_sheet_headers_cache
        if hasattr(self, '_sheet_headers_cache'):
            del self._sheet_headers_cache

        # 9. 强制垃圾回收
        gc.collect()
        self._log("已清理上一份文件的缓存数据")

    def _update_column_display(self):
        """根据当前文件状态更新列显示区域"""
        # 取消旧的Configure监听
        if hasattr(self, '_col_configure_id') and self._col_configure_id:
            self.col_area.unbind("<Configure>")
            self._col_configure_id = None

        # 清除旧的列单元格
        if self.column_cells_frame is not None:
            self.column_cells_frame.destroy()
            self.column_cells_frame = None
        self.column_cells.clear()

        file_path = self.file_path.get().strip()
        if not file_path or not os.path.exists(file_path):
            self.col_placeholder.configure(text="请先选择Excel文件")
            self.col_placeholder.pack(padx=10, pady=12)
            return

        # 读取表头
        headers = self._read_headers(file_path)
        if not headers:
            self.col_placeholder.configure(text="无法读取表头，请确认文件格式")
            self.col_placeholder.pack(padx=10, pady=12)
            return

        self.column_headers = headers
        self.col_placeholder.pack_forget()  # 隐藏占位文字

        # 缓存当前Sheet的表头（供智能切换使用）
        current_sheet = self.selected_sheet.get()
        if current_sheet:
            if not hasattr(self, '_sheet_headers_cache'):
                self._sheet_headers_cache = {}
            self._sheet_headers_cache[current_sheet] = tuple(headers)

        # 创建列单元格容器
        self.column_cells_frame = ctk.CTkFrame(self.col_area, fg_color="transparent")
        self.column_cells_frame.pack(fill="both", expand=True, padx=5, pady=8)

        # 先创建所有单元格（暂不布局）
        for idx, header in enumerate(headers):
            display = str(header).strip() if header else ""
            if not display:
                display = f"(空)"
            if len(display) > 6:
                display = display[:5] + "…"
            col_letter = get_column_letter(idx + 1)
            text = f"{col_letter}\n{display}"

            outer = ctk.CTkFrame(self.column_cells_frame, corner_radius=6,
                                 cursor="hand2", border_width=0,
                                 fg_color=COLOR_BG)
            inner = ctk.CTkFrame(outer, corner_radius=4, border_width=0,
                                 fg_color=COLOR_BG_WHITE, border_color=COLOR_BORDER)
            label = ctk.CTkLabel(inner, text=text,
                                 font=ctk.CTkFont(size=FONT_SIZE_SMALL, family=FONT_FAMILY),
                                 text_color=COLOR_TEXT_PRIMARY,
                                 width=75, height=42, anchor="center")
            label.pack(padx=4, pady=4)
            inner.pack(padx=3, pady=3)

            for widget in (outer, inner, label):
                widget.bind("<Button-1>", lambda e, i=idx: self._on_column_click(i))

            self.column_cells[idx] = (outer, inner)

        # 初始布局 + 监听宽度变化自动重排
        self.root.after_idle(self._relayout_columns)
        self._col_configure_id = self.col_area.bind(
            "<Configure>", lambda e: self.root.after_idle(self._relayout_columns)
        )

        # 应用已有选择状态
        self._redraw_column_cells()
        self._update_selection_info()

    def _smart_update_column_display(self):
        """智能更新列显示：仅当表头变化时才重建UI，否则只刷新选择状态（零IO）"""
        current_sheet = self.selected_sheet.get()
        file_path = self.file_path.get().strip()

        if not file_path or not os.path.exists(file_path):
            self._update_column_display()
            return

        # 直接从缓存取表头，不读文件（零IO）
        cached_headers = getattr(self, '_all_sheet_headers_cache', {})
        if current_sheet not in cached_headers:
            # 该Sheet从未缓存过，需要读取一次
            headers = self._read_headers(file_path)
            if not headers:
                self._update_column_display()
                return
            cached_headers[current_sheet] = tuple(headers)
            self._all_sheet_headers_cache = cached_headers
        headers_tuple = cached_headers[current_sheet]

        # 检查是否需要重建UI
        need_rebuild = (
            self.column_cells_frame is None or
            len(self.column_headers) != len(headers_tuple) or
            tuple(self.column_headers) != headers_tuple
        )

        if need_rebuild:
            # 表头变了或首次显示，完整重建（使用缓存的表头，不再读文件）
            self._init_column_display_with_headers(list(headers_tuple))
        else:
            # 表头没变，纯内存刷新选择状态（毫秒级）
            self._redraw_column_cells()
            self._update_selection_info()

    def _relayout_columns(self):
        """根据容器实际宽度重新排列列单元格的grid布局"""
        if not self.column_cells_frame or not self.column_cells:
            return

        # 取消上一个防抖回调
        if hasattr(self, '_relayout_after_id') and self._relayout_after_id:
            self.root.after_cancel(self._relayout_after_id)

        self._relayout_after_id = self.root.after(50, self._do_relayout_columns)

    def _do_relayout_columns(self):
        """实际执行列重排"""
        self._relayout_after_id = None
        if not self.column_cells_frame or not self.column_cells:
            return

        # 每个单元格宽度约 88px (75 + 3*2 + 3*2)，加间距 padx=3*2
        cell_width = 96
        try:
            available = self.column_cells_frame.winfo_width()
        except Exception:
            return
        if available <= 1:
            available = 400  # 默认值（窗口尚未渲染时）

        cols_per_row = max(1, available // cell_width)

        for idx, (outer, inner) in self.column_cells.items():
            outer.grid_forget()
            r = idx // cols_per_row
            c = idx % cols_per_row
            outer.grid(row=r, column=c, padx=3, pady=3)

    # ==================== 规则管理方法 ====================

    def _open_auto_select_dialog(self):
        """打开自动选择弹窗"""
        file_path = self.file_path.get()
        if not file_path or not os.path.exists(file_path):
            messagebox.showwarning("提示", "请先选择Excel文件！")
            return
        if not hasattr(self, '_all_sheet_headers_cache') or not self._all_sheet_headers_cache:
            messagebox.showwarning("提示", "请先加载Excel文件！")
            return

        dialog = AutoSelectDialog(self.root, file_path, self._all_sheet_headers_cache)
        dialog.dialog.wait_window()

        # 弹窗关闭后，检查是否确认了结果
        if hasattr(dialog, 'result') and dialog.result:
            # 将自动选择的规则导入到主页面规则列表
            for rule_data in dialog.result:
                rule = {
                    'green_selections': rule_data['green_selections'],
                    'red_selections': rule_data['red_selections']
                }
                self.saved_rules.append(rule)
            self._refresh_rules_list()
            sheet_names = [r['sheet_name'] for r in dialog.result]
            self._log(f"自动选择已导入 {len(dialog.result)} 条规则: {', '.join(sheet_names)}")

    def _save_rule(self):
        """保存当前选择的列为一条规则(仅当前Sheet)"""
        current_sheet = self.selected_sheet.get()
        if not current_sheet:
            messagebox.showwarning("提示", "请先选择Sheet页！")
            return
        
        if not self.green_selections:
            messagebox.showwarning("提示", "请先在当前Sheet上选择整体范围！")
            return
        if not self.red_selections:
            messagebox.showwarning("提示", "请先在当前Sheet上选择审查的列！")
            return
        
        # 检查是否在当前sheet上有选择
        if current_sheet not in self.green_selections or not self.green_selections[current_sheet]:
            messagebox.showwarning("提示", f"Sheet「{current_sheet}」上没有选择整体范围！")
            return
        if current_sheet not in self.red_selections or not self.red_selections[current_sheet]:
            messagebox.showwarning("提示", f"Sheet「{current_sheet}」上没有选择审查的列！")
            return

        # 只保存当前sheet的选择
        rule = {
            'green_selections': {current_sheet: copy.deepcopy(self.green_selections[current_sheet])},
            'red_selections': {current_sheet: copy.deepcopy(self.red_selections[current_sheet])}
        }
        self.saved_rules.append(rule)

        # 重置当前列选择
        self.green_selections.clear()
        self.red_selections.clear()
        self.selection_mode = None
        self._update_mode_buttons()
        self._redraw_column_cells()
        self._update_selection_info()

        # 刷新规则列表UI
        self._refresh_rules_list()
        self._log(f"已保存规则{len(self.saved_rules)}: Sheet「{current_sheet}」")

    def _delete_rule(self, rule_idx):
        """删除指定索引的规则"""
        if 0 <= rule_idx < len(self.saved_rules):
            self.saved_rules.pop(rule_idx)
            self._refresh_rules_list()

    def _clear_all_rules(self):
        """清空所有已保存的规则"""
        if not self.saved_rules:
            messagebox.showinfo("提示", "规则列表已经是空的！")
            return
        
        if messagebox.askyesno("确认清空", f"确定要清空所有规则吗？\n当前共有 {len(self.saved_rules)} 条规则。"):
            self.saved_rules.clear()
            self._refresh_rules_list()
            self._log("已清空所有规则")

    def _edit_rule(self, rule_idx):
        """编辑指定索引的规则(恢复到当前选择状态)"""
        if 0 <= rule_idx < len(self.saved_rules):
            rule = self.saved_rules[rule_idx]
            
            # 获取规则中的sheet(应该只有一个)
            green_sheets = list(rule['green_selections'].keys())
            red_sheets = list(rule['red_selections'].keys())
            
            if green_sheets:
                target_sheet = green_sheets[0]
            elif red_sheets:
                target_sheet = red_sheets[0]
            else:
                return
            
            # 切换到对应的sheet
            if target_sheet in self.sheet_names:
                self.selected_sheet.set(target_sheet)
                self.sheet_optionmenu.set(target_sheet)
            
            # 先清空当前选择
            self.green_selections.clear()
            self.red_selections.clear()
            
            # 从规则中恢复选择状态
            self.green_selections = copy.deepcopy(rule['green_selections'])
            self.red_selections = copy.deepcopy(rule['red_selections'])
            
            # 删除原规则
            self.saved_rules.pop(rule_idx)
            
            # 更新UI显示
            self._update_mode_buttons()
            self._update_column_display()  # 重新加载列显示
            self._update_selection_info()
            self._refresh_rules_list()
            
            self._log(f"已加载规则{rule_idx + 1}到编辑状态(Sheet「{target_sheet}」),修改后请重新保存")

    def _refresh_rules_list(self):
        """刷新规则列表的UI显示(卡片式布局)"""
        # 销毁所有旧的规则行
        for widget in self.rule_widgets:
            widget.destroy()
        self.rule_widgets.clear()
        # 清空旧的规则操作按钮引用
        if hasattr(self, '_rule_action_buttons'):
            self._rule_action_buttons.clear()

        if not self.saved_rules:
            self.rules_placeholder.pack(pady=8)
            return

        self.rules_placeholder.pack_forget()

        for i, rule in enumerate(self.saved_rules):
            # 创建卡片框架
            card_frame = ctk.CTkFrame(
                self.rules_list_frame,
                border_width=1, border_color=COLOR_BORDER,
                corner_radius=8, fg_color=COLOR_BG_WHITE
            )
            card_frame.pack(fill="x", pady=(5, 8), padx=5)
            
            # 卡片内部容器
            card_inner = ctk.CTkFrame(card_frame, fg_color="transparent")
            card_inner.pack(fill="both", expand=True, padx=8, pady=8)
            card_inner.grid_columnconfigure(0, weight=1)
            card_inner.grid_columnconfigure(1, weight=1)
            
            # 标题行
            title_row = ctk.CTkFrame(card_inner, fg_color="transparent")
            title_row.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 6))
            title_row.grid_columnconfigure(0, weight=1)
            
            title_label = ctk.CTkLabel(
                title_row, text=f"规则 {i + 1}",
                font=ctk.CTkFont(size=FONT_SIZE_BODY, weight="bold", family=FONT_FAMILY),
                text_color=COLOR_TEXT_PRIMARY, anchor="w"
            )
            title_label.grid(row=0, column=0, sticky="w")
            
            # 按钮容器
            btn_frame = ctk.CTkFrame(title_row, fg_color="transparent")
            btn_frame.grid(row=0, column=1, sticky="e")
            
            edit_btn = ctk.CTkButton(
                btn_frame, text="编辑", width=55, height=26,
                font=ctk.CTkFont(size=FONT_SIZE_SMALL, family=FONT_FAMILY),
                fg_color=COLOR_BLUE, hover_color="#0D47A1", text_color="white",
                corner_radius=5,
                command=lambda idx=i: self._edit_rule(idx)
            )
            edit_btn.pack(side="left", padx=(0, 5))
            
            del_btn = ctk.CTkButton(
                btn_frame, text="删除", width=55, height=26,
                font=ctk.CTkFont(size=FONT_SIZE_SMALL, family=FONT_FAMILY),
                fg_color=COLOR_RED, hover_color="#B71C1C", text_color="white",
                corner_radius=5,
                command=lambda idx=i: self._delete_rule(idx)
            )
            del_btn.pack(side="left")

            # 保存规则操作按钮引用，用于处理期间禁用
            if not hasattr(self, '_rule_action_buttons'):
                self._rule_action_buttons = []
            self._rule_action_buttons.append(edit_btn)
            self._rule_action_buttons.append(del_btn)
            
            # 绿色区域:整体范围
            green_frame = ctk.CTkFrame(
                card_inner,
                fg_color=COLOR_GREEN_LIGHT,
                corner_radius=6
            )
            green_frame.grid(row=1, column=0, sticky="ew", pady=(0, 4), padx=(0, 4))
            
            green_inner = ctk.CTkFrame(green_frame, fg_color="transparent")
            green_inner.pack(fill="both", expand=True, padx=8, pady=6)
            green_inner.grid_columnconfigure(0, weight=1)
            
            green_title = ctk.CTkLabel(
                green_inner, text="整体范围",
                font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, weight="bold", family=FONT_FAMILY),
                text_color=COLOR_TEXT_PRIMARY, anchor="w"
            )
            green_title.grid(row=0, column=0, sticky="w", pady=(0, 3))
            
            # 构建整体范围内容
            green_content_lines = []
            for sheet_name in sorted(rule['green_selections']):
                cols = sorted(rule['green_selections'][sheet_name])
                letters = [get_column_letter(c + 1) for c in cols]
                green_content_lines.append(f"  {sheet_name}: {', '.join(letters)}")
            
            green_content = ctk.CTkLabel(
                green_inner, text="\n".join(green_content_lines),
                font=ctk.CTkFont(size=FONT_SIZE_SMALL, family=FONT_FAMILY),
                text_color=COLOR_TEXT_MUTED, anchor="w", justify="left"
            )
            green_content.grid(row=1, column=0, sticky="w")
            
            # 红色区域:审查列
            red_frame = ctk.CTkFrame(
                card_inner,
                fg_color=COLOR_RED_LIGHT,
                corner_radius=6
            )
            red_frame.grid(row=1, column=1, sticky="ew", pady=(0, 4), padx=(4, 0))
            
            red_inner = ctk.CTkFrame(red_frame, fg_color="transparent")
            red_inner.pack(fill="both", expand=True, padx=8, pady=6)
            red_inner.grid_columnconfigure(0, weight=1)
            
            red_title = ctk.CTkLabel(
                red_inner, text="审查列",
                font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, weight="bold", family=FONT_FAMILY),
                text_color=COLOR_TEXT_PRIMARY, anchor="w"
            )
            red_title.grid(row=0, column=0, sticky="w", pady=(0, 3))
            
            # 构建审查列内容
            red_content_lines = []
            for sheet_name in sorted(rule['red_selections']):
                cols = sorted(rule['red_selections'][sheet_name])
                letters = [get_column_letter(c + 1) for c in cols]
                red_content_lines.append(f"  {sheet_name}: {', '.join(letters)}")
            
            red_content = ctk.CTkLabel(
                red_inner, text="\n".join(red_content_lines),
                font=ctk.CTkFont(size=FONT_SIZE_SMALL, family=FONT_FAMILY),
                text_color=COLOR_TEXT_MUTED, anchor="w", justify="left"
            )
            red_content.grid(row=1, column=0, sticky="w")
            
            self.rule_widgets.append(card_frame)

    # ==================== 计时器方法 ====================

    def _start_timer(self):
        """启动计时器"""
        import time as _time
        self.process_start_time = _time.time()
        self.timer_running = True
        self._tick_timer()

    def _tick_timer(self):
        """每秒更新计时器显示"""
        if self._is_closing or not self.timer_running:
            return
        import time as _time
        elapsed = _time.time() - self.process_start_time
        elapsed_str = self._format_time(elapsed)
        try:
            self.timer_label.configure(text=f"已处理：{elapsed_str} | 剩余：--:--:--")
            self.timer_after_id = self.root.after(1000, self._tick_timer)
        except Exception:
            pass

    def _stop_timer(self, total_elapsed):
        """停止计时器并显示最终时间"""
        self.timer_running = False
        if self.timer_after_id:
            self.root.after_cancel(self.timer_after_id)
            self.timer_after_id = None
        elapsed_str = self._format_time(total_elapsed)
        self.timer_label.configure(text=f"已处理：{elapsed_str} | 剩余：00:00:00")

    def _reset_timer(self):
        """重置计时器显示"""
        self.timer_running = False
        if self.timer_after_id:
            self.root.after_cancel(self.timer_after_id)
            self.timer_after_id = None
        self.timer_label.configure(text="已处理：00:00:00 | 剩余：--:--:--")

    def _on_closing(self):
        """窗口关闭时的安全退出处理"""
        self._is_closing = True

        # 1. 取消所有 pending 的 after() 定时器，防止窗口销毁后回调触发
        try:
            if self.timer_after_id:
                self.root.after_cancel(self.timer_after_id)
                self.timer_after_id = None
        except Exception:
            pass
        try:
            if hasattr(self, '_col_configure_id') and self._col_configure_id:
                self.root.after_cancel(self._col_configure_id)
        except Exception:
            pass
        try:
            if hasattr(self, '_relayout_after_id') and self._relayout_after_id:
                self.root.after_cancel(self._relayout_after_id)
        except Exception:
            pass
        # 取消大文件对话框的倒计时定时器
        try:
            if hasattr(self, '_dialog_after_id') and self._dialog_after_id is not None:
                self.root.after_cancel(self._dialog_after_id)
                self._dialog_after_id = None
        except Exception:
            pass

        # 2. 安全销毁大文件对话框（如果正在显示）
        try:
            if hasattr(self, '_large_file_dialog') and self._large_file_dialog is not None:
                self._large_file_dialog.grab_release()
                self._large_file_dialog.destroy()
                self._large_file_dialog = None
        except Exception:
            pass

        # 3. 终止仍在运行的 Go 子进程
        if hasattr(self, '_go_process') and self._go_process is not None:
            try:
                self._go_process.terminate()
                self._go_process.wait(timeout=3)
            except Exception:
                try:
                    self._go_process.kill()
                except Exception:
                    pass
            self._go_process = None

        # 3. 销毁窗口
        self.root.destroy()

    @staticmethod
    def _format_time(seconds):
        """将秒数格式化为 HH:MM:SS"""
        h = int(seconds // 3600)
        m = int((seconds % 3600) // 60)
        s = int(seconds % 60)
        return f"{h:02d}:{m:02d}:{s:02d}"

    # ==================== 日志与文件选择 ====================

    def _log(self, message):
        """输出日志到界面 + 缓冲区"""
        if self._is_closing:
            return  # 窗口关闭中，跳过UI操作
        timestamp = time.strftime("%H:%M:%S")
        log_msg = f"[{timestamp}] {message}\n"
        self._log_buffer.append(log_msg)
        try:
            self.log_text.configure(state="normal")
            self.log_text.insert("end", log_msg)
            self.log_text.see("end")
            self.log_text.configure(state="disabled")
            self.root.update()
        except Exception:
            pass  # 窗口已销毁，静默忽略

    def _export_log_file(self):
        """将本次处理的完整日志导出为文件"""
        if not self._log_buffer:
            return
        try:
            # 日志文件放在源文件同目录下，命名为：原文件名_processing_log.txt
            src_name = os.path.splitext(os.path.basename(self._processing_file_path))[0]
            log_dir = os.path.dirname(self._processing_file_path) or "."
            log_path = os.path.join(log_dir, f"{src_name}_processing_log.txt")

            with open(log_path, 'w', encoding='utf-8') as f:
                f.write("=" * 70 + "\n")
                f.write(f"Excel数据清洗工具 - 处理日志\n")
                f.write("=" * 70 + "\n")
                f.write(f"生成时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"源文件: {self._processing_file_path}\n")
                f.write(f"规则模式: {self.rule_mode_var.get()}\n")
                f.write(f"跳过表头: {self.skip_header_var.get()}\n")
                f.write(f"已保存规则数: {len(self.saved_rules)}\n")
                if self.saved_rules:
                    f.write("-" * 50 + "\n")
                    for i, rule in enumerate(self.saved_rules):
                        green = {}
                        for sn, cols in rule['green_selections'].items():
                            green[sn] = [get_column_letter(c + 1) for c in sorted(cols)]
                        red = {}
                        for sn, cols in rule['red_selections'].items():
                            red[sn] = [get_column_letter(c + 1) for c in sorted(cols)]
                        f.write(f"  规则{i+1}: 整体范围={green}, 审查列={red}\n")
                f.write("=" * 70 + "\n\n")

                f.write("--- Python日志 ---\n")
                f.writelines(self._log_buffer)
                f.write("\n")

                if self._go_log_content:
                    f.write("--- Go引擎日志 ---\n")
                    f.write(self._go_log_content)
                    f.write("\n")

            self._log(f"✓ 处理日志已导出: {log_path}")
        except Exception as e:
            self._log(f"⚠ 导出处理日志失败: {e}")

    def _select_file(self):
        """选择Excel文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if file_path:
            # 切换文件前清理旧文件状态
            self._cleanup_file_state()
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            self.file_path.set(file_path)
            self._log(f"已选择文件：{file_path}（{file_size_mb:.1f} MB）")

            if file_size_mb > 10:
                self._show_large_file_dialog(file_path)
            else:
                self._load_file_and_init_ui(file_path)

    def _show_large_file_dialog(self, file_path):
        """大文件提醒弹窗：后台加载+动态倒计时+可取消"""
        dialog = ctk.CTkToplevel(self.root)
        dialog.title("提醒")
        dialog.geometry("460x260")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.attributes("-topmost", True)

        # 加载过程中禁用关闭按钮
        dialog.protocol("WM_DELETE_WINDOW", lambda: None)

        # 保存对话框引用，供关闭时安全销毁
        self._large_file_dialog = dialog
        self._dialog_after_id = None  # 保存倒计时after_id，供取消用

        # 居中
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 460) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 260) // 2
        dialog.geometry(f"+{x}+{y}")

        # 内容容器
        content_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, padx=20)

        # 提示文字
        msg_label = ctk.CTkLabel(
            content_frame, text="文件数据量过大，正在后台加载...",
            font=ctk.CTkFont(size=FONT_SIZE_BODY, family=FONT_FAMILY),
            text_color=COLOR_TEXT_PRIMARY, wraplength=400
        )
        msg_label.pack(pady=(15, 3))

        # 倒计时标签（加粗深色）
        countdown_label = ctk.CTkLabel(
            content_frame, text="加载中，请稍候...",
            font=ctk.CTkFont(size=FONT_SIZE_BODY + 2, weight="bold", family=FONT_FAMILY),
            text_color=COLOR_TEXT_PRIMARY
        )
        countdown_label.pack(pady=3)

        # 左下角提示（默认隐藏，加载完成后显示）
        tip_label = ctk.CTkLabel(
            content_frame, text="提示：处理过程中可能会出现 无响应/卡死 情况，属于正常现象，请耐心等待！",
            font=ctk.CTkFont(size=FONT_SIZE_SMALL + 1, family=FONT_FAMILY),
            text_color=COLOR_RED,
            wraplength=420, anchor="w", justify="left"
        )
        tip_label.pack(side="bottom", fill="x", pady=(5, 0))
        tip_label.pack_forget()

        # 按钮行容器
        btn_row = ctk.CTkFrame(content_frame, fg_color="transparent")
        btn_row.pack(pady=(5, 10))

        # 取消按钮（始终可用）
        cancel_btn = ctk.CTkButton(
            btn_row, text="取消",
            font=ctk.CTkFont(size=FONT_SIZE_BODY, family=FONT_FAMILY),
            width=80, height=36,
            fg_color="#757575", hover_color="#616161", text_color="white",
            corner_radius=6,
            command=self._on_dialog_cancel
        )
        cancel_btn.pack(side="right", padx=(15, 0))

        # 继续按钮（初始置灰）
        continue_btn = ctk.CTkButton(
            btn_row, text="继续",
            font=ctk.CTkFont(size=FONT_SIZE_BODY, family=FONT_FAMILY),
            width=80, height=36,
            fg_color=COLOR_BLUE, hover_color="#0D47A1", text_color="white",
            corner_radius=6,
            state="disabled", command=lambda: self._on_dialog_continue(file_path, dialog)
        )
        continue_btn.pack(side="right")

        # 标记是否已取消/完成
        _dialog_done = [False]  # 用列表以便闭包内修改

        # 后台加载：只做纯IO和数据计算，不操作任何UI
        def _do_load():
            load_start = time.time()
            sheet_names = []
            headers = None
            load_error = None
            # 预读所有Sheet的表头（一次IO，后续切换零延迟）
            all_headers_cache = {}
            try:
                wb = load_workbook(filename=file_path, read_only=True)
                sheet_names = list(wb.sheetnames)
                # 检查空文件
                if not sheet_names:
                    raise EmptyFileError(file_path)
                # 预读所有Sheet的表头（workbook已打开，几乎零额外开销）
                for sn in sheet_names:
                    if sn in wb.sheetnames:
                        ws = wb[sn]
                        for row in ws.iter_rows(max_row=1, values_only=True):
                            if row:
                                all_headers_cache[sn] = tuple(row)
                            break
                # 第一个Sheet的表头作为默认显示
                headers = list(all_headers_cache.get(sheet_names[0], [])) if sheet_names else None
                wb.close()
            except InvalidFileException as e:
                biz_err, _ = self._classify_error(e, context="加载大文件Excel", file_path=file_path)
                load_error = f"{biz_err.error_code}: {biz_err.message}"
            except PermissionError as e:
                biz_err, _ = self._classify_error(e, context="打开大文件", file_path=file_path, is_locked=True)
                load_error = f"{biz_err.error_code}: {biz_err.message}"
            except BadZipFile as e:
                biz_err, _ = self._classify_error(e, context="解析大文件", file_path=file_path)
                load_error = f"{biz_err.error_code}: {biz_err.message}"
            except MemoryError as e:
                file_mb = os.path.getsize(file_path) / (1024*1024) if os.path.exists(file_path) else 0
                biz_err = MemoryError_(file_size_mb=file_mb)
                load_error = f"{biz_err.error_code}: {biz_err.message}"
            except Exception as e:
                biz_err, detail = self._classify_error(e, context="加载大文件", file_path=file_path)
                load_error = f"{biz_err.error_code}: {detail}"
            load_time = time.time() - load_start

            # 存储结果（纯数据赋值，线程安全）
            self._pending_file_result = (sheet_names, headers, load_error, all_headers_cache)

            # 加载完成立即回到主线程激活按钮（无需等待倒计时）
            if not self._is_closing:
                self.root.after(0, _on_load_complete, load_time)

        def _safe_configure(widget, **kwargs):
            """安全configure，窗口销毁后静默忽略"""
            try:
                widget.configure(**kwargs)
            except Exception:
                pass

        def _on_load_complete(load_time):
            """主线程：加载完成，立即激活继续按钮（无需等待）"""
            if self._is_closing or _dialog_done[0]:
                return
            try:
                msg_label.configure(text=f'加载完成（耗时 {load_time:.1f}s），请点击【继续】')
                countdown_label.configure(text="")
                continue_btn.configure(state="normal")
                cancel_btn.pack_forget()  # 完成后隐藏取消按钮
                tip_label.pack(side="bottom", fill="x", pady=(5, 0))
                # 加载完成，关闭按钮等同于点击"继续"
                dialog.protocol("WM_DELETE_WINDOW", lambda: self._on_dialog_continue(file_path, dialog))
                self._dialog_after_id = None
            except Exception:
                pass

        threading.Thread(target=_do_load, daemon=True).start()

    def _on_dialog_cancel(self):
        """大文件对话框取消按钮：清理状态并销毁对话框"""
        # 取消倒计时定时器
        if hasattr(self, '_dialog_after_id') and self._dialog_after_id is not None:
            try:
                self.root.after_cancel(self._dialog_after_id)
            except Exception:
                pass
            self._dialog_after_id = None
        # 清理对话框
        if hasattr(self, '_large_file_dialog') and self._large_file_dialog is not None:
            try:
                self._large_file_dialog.destroy()
            except Exception:
                pass
            self._large_file_dialog = None
        # 清理待定结果
        self._pending_file_result = ([], None, "用户取消了文件加载", {})

    def _on_dialog_continue(self, file_path, dialog):
        """大文件弹窗点击继续后的初始化（数据已在后台预加载，无需再读文件）"""
        dialog.destroy()
        sheet_names, headers, load_error, all_headers_cache = self._pending_file_result

        if load_error:
            self.sheet_names = []
            self._log(f"文件加载失败：{load_error}")
        else:
            self.sheet_names = sheet_names
            # 缓存所有Sheet表头，后续切换零IO
            if all_headers_cache:
                self._all_sheet_headers_cache = all_headers_cache

        # 更新Sheet选择器
        if self.sheet_names:
            self.selected_sheet.set(self.sheet_names[0])
            self.sheet_optionmenu.configure(values=self.sheet_names)
            self.sheet_row.grid()
        else:
            self.sheet_row.grid_remove()

        # 重置模式按钮
        self._update_mode_buttons()

        # 使用后台预读的表头直接初始化列显示，不再二次打开文件
        if headers:
            self._init_column_display_with_headers(headers)
        else:
            self._update_column_display()

    def _load_file_and_init_ui(self, file_path):
        """小文件直接加载并初始化UI（预读所有Sheet表头）"""
        all_headers_cache = {}
        try:
            wb = load_workbook(filename=file_path, read_only=True)
            self.sheet_names = list(wb.sheetnames)
            # 检查空文件
            if not self.sheet_names:
                wb.close()
                raise EmptyFileError(file_path)
            # 预读所有Sheet的表头（workbook已打开，几乎零额外开销）
            for sn in self.sheet_names:
                ws = wb[sn]
                for row in ws.iter_rows(max_row=1, values_only=True):
                    if row:
                        all_headers_cache[sn] = tuple(row)
                    break
            wb.close()
            # 缓存所有Sheet表头，后续切换零IO
            self._all_sheet_headers_cache = all_headers_cache
        except EmptyFileError as e:
            self._safe_log_error(e)
            messagebox.showerror(f"错误 {e.error_code}", e.get_user_message())
            self.sheet_names = []
        except InvalidFileException as e:
            biz_err, detail = self._classify_error(e, context="加载Excel", file_path=file_path)
            self._safe_log_error(biz_err, detail)
            messagebox.showerror(f"错误 {biz_err.error_code}", biz_err.get_user_message())
            self.sheet_names = []
        except PermissionError as e:
            biz_err, detail = self._classify_error(e, context="打开文件", file_path=file_path)
            self._safe_log_error(biz_err, detail)
            messagebox.showerror(f"错误 {biz_err.error_code}", biz_err.get_user_message())
            self.sheet_names = []
        except BadZipFile as e:
            biz_err, detail = self._classify_error(e, context="解析Excel", file_path=file_path)
            self._safe_log_error(biz_err, detail)
            messagebox.showerror(f"错误 {biz_err.error_code}", biz_err.get_user_message())
            self.sheet_names = []
        except Exception as e:
            biz_err, detail = self._classify_error(e, context="加载文件", file_path=file_path)
            self._safe_log_error(biz_err, detail)
            self.sheet_names = []

        # 更新Sheet选择器
        if self.sheet_names:
            self.selected_sheet.set(self.sheet_names[0])
            self.sheet_optionmenu.configure(values=self.sheet_names)
            self.sheet_row.grid()  # 显示Sheet选择行
        else:
            self.sheet_row.grid_remove()  # 隐藏

        # 重置模式按钮
        self._update_mode_buttons()
        # 更新列显示
        self._update_column_display()


    # ==================== 处理控制 ====================

    def _toggle_pause(self):
        """切换暂停/继续状态"""
        # 检查是否在Go全托管模式（Go是独立外部进程，无法被Python的pause_event控制）
        go_proc = getattr(self, '_go_process', None)
        if go_proc is not None and go_proc.poll() is None:
            # Go进程正在运行 → 显示提示而非真正暂停
            messagebox.showinfo(
                "提示",
                "当前使用 Go 全托管模式处理数据。\n\n"
                "由于 Go 程序是独立运行的子进程，\n"
                "暂不支持暂停操作。\n\n"
                "如需停止处理，请点击「取消」按钮。"
            )
            return

        if self._pause_event.is_set():
            # 当前运行中 → 暂停
            self._pause_event.clear()
            self.pause_btn.configure(text="▶ 继续", fg_color="#2E7D32", hover_color="#388E3C")
            self._log("⏸ 处理已暂停...")
        else:
            # 当前暂停中 → 继续
            self._pause_event.set()
            self.pause_btn.configure(text="⏸ 暂停", fg_color="#F57F17", hover_color="#F9A825")
            self._log("▶ 处理已继续...")

    def _cancel_processing(self):
        """取消处理"""
        if messagebox.askyesno("确认取消", "确定要取消当前处理吗？\n已处理的数据将不会保存。"):
            self._cancel_flag = True
            self._pause_event.set()  # 解除暂停以便线程能检测到取消

            # 杀掉正在运行的Go子进程（如果存在）
            go_proc = getattr(self, '_go_process', None)
            if go_proc is not None and go_proc.poll() is None:
                try:
                    go_proc.kill()
                    go_proc.wait(timeout=5)
                    self._log("✖ 已终止Go子进程")
                except Exception as e:
                    self._log(f"⚠ 终止Go子进程失败: {e}")

            self._log("✖ 用户已取消处理...")

    # ==================== 处理逻辑 ====================

    def _start_processing(self):
        """启动处理流程"""
        file_path = self.file_path.get().strip()

        if not file_path:
            messagebox.showwarning("提示", "请先选择要处理的Excel文件！")
            return

        if not os.path.exists(file_path):
            messagebox.showerror("错误", f"文件不存在：{file_path}")
            return

        if not file_path.lower().endswith('.xlsx'):
            messagebox.showwarning("提示", "请选择.xlsx格式的Excel文件！")
            return

        if not self.saved_rules:
            messagebox.showwarning("提示", "请先选择列并保存至少一条规则！")
            return

        # 初始化控制标志
        self._pause_event.set()
        self._cancel_flag = False

        # 记录当前处理的文件（用于日志导出）
        self._processing_file_path = file_path
        self._log_buffer.clear()
        self._go_log_content = ""

        # 临时诊断：启用关键词追踪（定位跨Sheet去重失败原因）
        self._trace_keywords = ["北京华医圣杰"]

        # 打印所有规则
        for i, rule in enumerate(self.saved_rules):
            green_parts = []
            for sn in sorted(rule['green_selections']):
                letters = [get_column_letter(c + 1) for c in sorted(rule['green_selections'][sn])]
                green_parts.append(f"{sn}+{','.join(letters)}")
            red_parts = []
            for sn in sorted(rule['red_selections']):
                letters = [get_column_letter(c + 1) for c in sorted(rule['red_selections'][sn])]
                red_parts.append(f"{sn}+{','.join(letters)}")
            self._log(f"规则{i + 1} → 整体范围：[{', '.join(green_parts)}]，审查列：[{', '.join(red_parts)}]")

        # 切换按钮状态
        self.process_btn.configure(state="disabled", text="⏳ 处理中...")
        self.pause_btn.configure(state="normal")
        self.cancel_btn.configure(state="normal")
        # 禁用左侧配置区域的所有控件（底部三按钮和右侧日志保持可用）
        self._lock_config_ui()
        self._start_timer()
        thread = threading.Thread(target=self._process_excel_with_rules, args=(file_path,))
        thread.daemon = True
        thread.start()

    # ==================== 底层操作（Go子程序 / XML备用） ====================

    def _get_shifter_exe_path(self):
        """
        获取Go处理程序的路径。
        优先级：
          1. PyInstaller打包后，从 sys._MEIPASS 读取
          2. 同目录下的 xlsx_shifter.exe（开发调试/手动部署）
          3. xlsx_shifter 子目录下的 exe
        """
        # 1. PyInstaller打包模式
        if getattr(sys, 'frozen', False):
            meipass = sys._MEIPASS
            meipass_path = os.path.join(meipass, "xlsx_shifter.exe")
            if os.path.isfile(meipass_path):
                return meipass_path

        # 2. 脚本同目录
        base_dir = os.path.dirname(os.path.abspath(__file__))

        candidates = [
            os.path.join(base_dir, "xlsx_shifter.exe"),       # 打包后在同目录
            os.path.join(base_dir, "xlsx_shifter", "xlsx_shifter.exe"),  # 开发时在子目录
            os.path.join(base_dir, "xlsx_shifter", "xlsx_shifter"),      # Linux/Mac无扩展名
        ]

        for path in candidates:
            if os.path.isfile(path):
                return path

        return None

    def _shift_up_go(self, xlsx_path, sheet_name, dup_row_indices, col_indices):
        """
        通过Go子程序执行上移操作（推荐方式，内存占用最低）。
        
        原理：调用编译好的xlsx_shifter.exe，通过JSON传递参数，
              Go程序内部使用excellize库直接操作xlsx，内存占用极低。
        
        参数:
            xlsx_path: xlsx文件路径
            sheet_name: Sheet名称
            dup_row_indices: 需要删除的行号列表(0-based)，已按降序排列
            col_indices: 需要上移的列索引列表(0-based)
        """
        shifter_exe = self._get_shifter_exe_path()
        if not shifter_exe:
            raise FileNotFoundError(
                f"未找到xlsx_shifter处理程序。"
                f"\n请确认以下位置之一存在该文件：\n"
                f"  - {os.path.dirname(os.path.abspath(__file__))}\\xlsx_shifter.exe\n"
                f"  - {os.path.dirname(os.path.abspath(__file__))}\\xlsx_shifter\\xlsx_shifter.exe\n\n"
                f"如需重新编译，请在 xlsx_shifter 目录下执行：\n  go build -o xlsx_shifter.exe main.go"
            )

        # 构建JSON任务输入
        task_input = {
            "file_path": xlsx_path,
            "sheet_name": sheet_name,
            "col_indices": list(col_indices),
            "dup_rows": list(dup_row_indices),  # 已降序排列
        }

        # 调用Go子进程
        self._log(f"    调用Go处理程序: {len(dup_row_indices)} 行 × {len(col_indices)} 列")

        try:
            process = subprocess.run(
                [shifter_exe, "--json"],
                input=json.dumps(task_input, ensure_ascii=False).encode("utf-8"),
                capture_output=True,
                timeout=3600,  # 最长1小时超时
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0,
            )
        except subprocess.TimeoutExpired:
            raise MemoryError_(file_size_mb=0)  # 超时当内存问题处理
        except FileNotFoundError as e:
            raise FileNotFoundError_(
                f"无法启动处理程序：{shifter_exe}\n请确保文件存在且可执行。"
            )

        # 解析结果
        try:
            result = json.loads(process.stdout.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            error_msg = process.stderr.decode("utf-8", errors="replace")[:500]
            raise WriteFileError(xlsx_path, operation="Go处理", detail=f"解析结果失败，原始输出：{error_msg}")

        if not result.get("success", False):
            err_msg = result.get("error_message", "未知错误")
            # 根据错误信息映射到业务异常
            if "不存在" in err_msg and "工作表" in err_msg:
                raise SheetNotFoundError(sheet_name)
            elif "打开文件" in err_msg:
                raise FileCorruptError(xlsx_path, detail=err_msg)
            elif "保存" in err_msg:
                raise WriteFileError(xlsx_path, operation="保存", detail=err_msg)
            else:
                raise WriteFileError(xlsx_path, operation="Go处理", detail=err_msg)

        modified = result.get("modified_rows", 0)
        cells_changed = result.get("cells_changed", 0)

        self._log(f"    [{sheet_name}] Go处理完成: {modified} 行, {cells_changed} 个单元格")

    def _shift_up_fallback(self, xlsx_path, sheet_name, dup_row_indices, col_indices):
        """
        备用方案：纯Python ElementTree操作XML（不依赖Go）。
        
        当Go程序不可用时自动降级到此方案。
        内存占用高于Go方案，但比openpyxl写模式低很多。
        """
        target_col_letters = {get_column_letter(c + 1) for c in col_indices}

        sheet_xml_name = None
        with zipfile.ZipFile(xlsx_path, 'r') as zf:
            try:
                with zf.open('xl/workbook.xml') as f:
                    wb_tree = ET.parse(f)
                    wb_root = wb_tree.getroot()
                    for sheet_elem in wb_root.iter('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'):
                        if sheet_elem.get('name') == sheet_name:
                            r_id = sheet_elem.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                            break
                    else:
                        raise SheetNotFoundError(sheet_name)
            except ET.ParseError as e:
                raise FileCorruptError(xlsx_path, detail=f"workbook.xml解析失败：{str(e)}")

            try:
                with zf.open('xl/_rels/workbook.xml.rels') as f:
                    rels_tree = ET.parse(f)
                    rels_root = rels_tree.getroot()
                    for rel in rels_root:
                        if rel.get('Id') == r_id:
                            target_path = rel.get('Target')
                            break
                    else:
                        raise SheetNotFoundError(sheet_name)
                sheet_xml_name = 'xl/' + target_path
            except (ET.ParseError, UnboundLocalError) as e:
                raise FileCorruptError(xlsx_path, detail=f"关系文件解析失败：{str(e)}")

        with zipfile.ZipFile(xlsx_path, 'r') as zf:
            if sheet_xml_name not in zf.namelist():
                raise SheetNotFoundError(sheet_name)
            xml_bytes = zf.read(sheet_xml_name)

        tree = ET.fromstring(xml_bytes)
        NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
        sheet_data = tree.find(f'{NS}sheetData')
        if sheet_data is None:
            self._log(f"    ⚠ Sheet「{sheet_name}」没有数据区域")
            return

        all_rows = list(sheet_data.findall(f'{NS}row'))
        row_map = {}
        max_row_num = 0
        for row_elem in all_rows:
            rn = int(row_elem.get('r', 0))
            row_map[rn] = row_elem
            if rn > max_row_num:
                max_row_num = rn

        cells_modified = 0

        for dup_idx, row_idx_0based in enumerate(dup_row_indices):
            self._pause_event.wait()
            if self._cancel_flag:
                return

            dup_row_1based = row_idx_0based + 1

            for move_from_row in range(dup_row_1based + 1, max_row_num + 1):
                from_row_elem = row_map.get(move_from_row)
                to_row_elem = row_map.get(move_from_row - 1)

                if from_row_elem is None:
                    continue

                from_cells_to_move = []
                for cell in from_row_elem.findall(f'{NS}c'):
                    cell_ref = cell.get('r', '')
                    col_letter = ''.join(c for c in cell_ref if c.isalpha())
                    if col_letter in target_col_letters:
                        from_cells_to_move.append(cell)

                if not from_cells_to_move:
                    continue

                if to_row_elem is None:
                    to_row_elem = ET.SubElement(sheet_data, f'{NS}row',
                                                  attrib={'r': str(move_from_row - 1)})
                    row_map[move_from_row - 1] = to_row_elem

                existing_cells = []
                for cell in to_row_elem.findall(f'{NS}c'):
                    cell_ref = cell.get('r', '')
                    col_letter = ''.join(c for c in cell_ref if c.isalpha())
                    if col_letter in target_col_letters:
                        existing_cells.append(cell)
                for ec in existing_cells:
                    to_row_elem.remove(ec)

                for cell in from_cells_to_move:
                    old_ref = cell.get('r', '')
                    col_letter = ''.join(c for c in old_ref if c.isalpha())
                    new_ref = f"{col_letter}{move_from_row - 1}"
                    cell.set('r', new_ref)
                    from_row_elem.remove(cell)
                    to_row_elem.append(cell)
                    cells_modified += 1

            last_row_elem = row_map.get(max_row_num)
            if last_row_elem is not None:
                cells_to_remove = []
                for cell in last_row_elem.findall(f'{NS}c'):
                    cell_ref = cell.get('r', '')
                    col_letter = ''.join(c for c in cell_ref if c.isalpha())
                    if col_letter in target_col_letters:
                        cells_to_remove.append(cell)
                for cr in cells_to_remove:
                    last_row_elem.remove(cr)

            if dup_idx > 0 and dup_idx % 500 == 0:
                gc.collect()

        new_xml_bytes = ET.tostring(tree, encoding='UTF-8', xml_declaration=True)

        temp_path = xlsx_path + '.tmp_shift'
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.namelist():
                    if item == sheet_xml_name:
                        zout.writestr(item, new_xml_bytes)
                    else:
                        zout.writestr(item, zin.read(item))

        os.replace(temp_path, xlsx_path)

        self._log(f"    [{sheet_name}] 备用方案完成: {len(dup_row_indices)} 行, {cells_modified} 个单元格")

    def _process_via_go_full(self, output_file):
        """
        Go全托管模式：将所有规则一次性传给Go，由Go完成扫描+上移。
        内存占用极低（<200MB），推荐用于大文件。
        """
        shifter_exe = self._get_shifter_exe_path()
        if not shifter_exe or not os.path.isfile(shifter_exe):
            return False  # Go不可用，返回False让调用方走Python路径

        rule_mode = self.rule_mode_var.get()
        skip_header = self.skip_header_var.get()

        # 构建规则列表
        rules = []
        for rule in self.saved_rules:
            green_sel = rule['green_selections']
            red_sel = rule['red_selections']
            involved = set(green_sel.keys()) & set(red_sel.keys())
            if not involved:
                continue
            for sheet_name in sorted(involved):
                rules.append({
                    "sheet_name": sheet_name,
                    "red_cols": sorted(red_sel[sheet_name]),
                    "green_cols": sorted(green_sel[sheet_name]),
                })

        if not rules:
            return True  # 无有效规则，直接成功

        task_input = {
            "mode": "full",
            "file_path": output_file,
            "rules": rules,
            "skip_header": skip_header,
            "rule_mode": rule_mode,
        }

        self._log(f"    ✓ Go全托管模式启动: {len(rules)} 条规则")
        self._log(f"    ✓ 处理程序: {shifter_exe}")
        self._log(f"    ⏳ 正在调用Go程序，请稍候...")

        import threading as _threading
        import queue as _queue

        import time as _time

        # ====== 方案：stderr 写入临时文件，主线程轮询读取 ======
        # 彻底解决：1) communicate() 吞 stderr 问题  2) 管道缓冲延迟
        # 注意：用二进制模式(wb)打开，因为Go直接写原始字节到fd，绕过Python文本编码层
        _log_file_path = os.path.join(os.path.dirname(__file__), f'.go_log_{os.getpid()}.tmp')

        try:
            with open(_log_file_path, 'wb') as _lf:
                pass  # 清空/创建日志文件

            # 构建环境变量（复制当前环境 + 追踪关键词）
            _env = os.environ.copy()
            if hasattr(self, '_trace_keywords') and self._trace_keywords:
                _env["DEDUP_TRACE_KEYWORDS"] = ",".join(self._trace_keywords)
                self._log(f"    [TRACE] 启用关键词追踪: {self._trace_keywords}")

            process = subprocess.Popen(
                [shifter_exe],
                stdin=subprocess.PIPE,
                stdout=subprocess.PIPE,
                stderr=open(_log_file_path, 'wb'),
                env=_env,
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0,
            )
            self._go_process = process  # 保存引用，供取消按钮使用

            # 写入任务JSON并关闭stdin，让Go开始处理
            task_json = json.dumps(task_input, ensure_ascii=False).encode("utf-8")
            process.stdin.write(task_json)
            process.stdin.close()

            # --- 共享容器 ---
            _comm_result = {'stdout': b'', 'error': None}
            _last_log_pos = 0  # 文件上次读取位置（游标）

            # --- 子线程：等待进程完成，只读 stdout（不再碰 stderr）---
            def _wait_for_process():
                try:
                    out, _ = process.communicate(timeout=7200)
                    _comm_result['stdout'] = out
                except subprocess.TimeoutExpired:
                    process.kill()
                    process.wait()
                    _comm_result['error'] = MemoryError_(file_size_mb=0)
                except (OSError, ValueError) as e:
                    if self._cancel_flag:
                        _comm_result['stdout'] = '{"success":false,"error_message":"user cancelled"}'.encode("ascii")
                    else:
                        _comm_result['error'] = e

            comm_thread = _threading.Thread(target=_wait_for_process, daemon=True)
            comm_thread.start()

            # --- 主线程：非阻塞等待循环 ---
            # 从日志文件读取增量内容显示 + root.update() 保持GUI响应
            while comm_thread.is_alive():
                # 从临时日志文件读取新增内容
                try:
                    with open(_log_file_path, 'rb') as _lf:
                        _lf.seek(_last_log_pos)
                        _new_data = _lf.read()
                        if _new_data:
                            _last_log_pos = _lf.tell()
                            _text = _new_data.decode('utf-8', errors='replace')
                            for _line in _text.splitlines():
                                _stripped = _line.strip()
                                if _stripped:
                                    self._log(f"      {_stripped}")
                except (FileNotFoundError, OSError):
                    pass

                if not self._is_closing:
                    self.root.update()
                _time.sleep(0.15)  # 150ms：减少文件IO频率

            # 进程已结束，读取最终日志
            try:
                with open(_log_file_path, 'rb') as _lf:
                    _lf.seek(_last_log_pos)
                    _new_data = _lf.read()
                    _text = _new_data.decode('utf-8', errors='replace')
                    for _line in _text.splitlines():
                        _stripped = _line.strip()
                        if _stripped:
                            self._log(f"      {_stripped}")
            except (FileNotFoundError, OSError):
                pass

            # 检查是否有异常
            if _comm_result['error'] is not None:
                raise _comm_result['error']

            stdout_data = _comm_result['stdout']

            # 如果进程被用户取消kill掉，主动抛出取消异常
            if self._cancel_flag and (process.returncode is None or process.returncode != 0):
                raise InterruptedError("处理已被用户取消")

        except FileNotFoundError:
            return False  # Go程序不存在，降级
        finally:
            self._go_process = None  # 清理引用
            # 保存Go日志内容到缓冲区（用于导出完整日志）
            try:
                if os.path.exists(_log_file_path):
                    with open(_log_file_path, 'rb') as _lf:
                        self._go_log_content = _lf.read().decode('utf-8', errors='replace')
                    os.remove(_log_file_path)
            except OSError:
                pass

        # 解析结果
        try:
            result = json.loads(stdout_data.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            # 优先使用已保存到内存的Go日志（finally块中已读取），其次读文件
            _err_ctx = "(无日志输出)"
            # 1. 先用内存中的日志（finally块已保存，文件可能已被删除）
            if hasattr(self, '_go_log_content') and self._go_log_content:
                _all = self._go_log_content.splitlines()
                if _all:
                    _err_ctx = "\n".join(l.strip() for l in _all[-10:] if l.strip())
            else:
                # 2. 内存也没有，尝试读磁盘（可能还存在）
                try:
                    with open(_log_file_path, 'rb') as _lf:
                        _raw = _lf.read()
                        _all = _raw.decode('utf-8', errors='replace').splitlines()
                        if _all:
                            _err_ctx = "\n".join(l.strip() for l in _all[-10:] if l.strip())
                except (FileNotFoundError, OSError):
                    pass
            # 3. 附加stdout原始内容前100字符帮助诊断
            _stdout_hint = ""
            try:
                _raw_stdout = stdout_data[:200].decode('utf-8', errors='replace')
                if _raw_stdout.strip():
                    _stdout_hint = f" | stdout前200字: [{_raw_stdout}]"
            except Exception:
                pass
            raise WriteFileError(output_file, operation="Go全托管",
                                 detail=f"解析结果失败: {_err_ctx[:500]}{_stdout_hint}")

        if not result.get("success", False):
            err_msg = result.get("error_message", "未知错误")
            raise WriteFileError(output_file, operation="Go全托管", detail=err_msg)

        # 显示结果摘要
        total_dups = result.get("total_dups", 0)
        memory_mb = result.get("memory_mb", 0)
        self._log(f"    ✓ Go处理完成: {total_dups} 个重复项已处理, 内存 {memory_mb:.1f}MB")

        rule_results = result.get("rule_results", [])
        for rr in rule_results:
            sn = rr.get("sheet_name", "?")
            dups = rr.get("total_dups", 0)
            tms = rr.get("time_ms", 0)
            self._log(f"      [{sn}] 发现 {dups} 个重复项, 耗时 {tms:.0f}ms")

        # 显示阶段2实际删除结果（JSON内嵌诊断）
        stage2_diag = result.get("stage2_diag", [])
        if stage2_diag:
            total_skipped = 0
            for d in stage2_diag:
                sn = d.get("sheet_name", "?")
                kept = d.get("kept_count", 0)
                skipped = d.get("skipped_count", 0)
                total = d.get("total_rows", 0)
                orig = d.get("original_size", 0)
                res = d.get("result_size", 0)
                total_skipped += skipped
                if skipped > 0:
                    self._log(f"      ⛔ [{sn}] 阶段2: 保留{kept}行/删除{skipped}重复, 原始行数{total}, 大小{orig}→{res}")
                else:
                    self._log(f"      ⚠ [{sn}] 阶段2: 保留{kept}行/删除0重复(未生效), 原始行数{total}")
            self._log(f"      📊 阶段2合计: 删除{total_skipped}个重复行")
        else:
            # 诊断：stage2_diag为空的可能原因
            has_diag_key = "stage2_diag" in result
            all_keys = list(result.keys())
            self._log(f"      ⚠ 阶段2诊断: stage2_diag{'存在但为空' if has_diag_key else '不存在(未设置)'}, JSON字段={all_keys}")

        return True

    def _shift_up_lowlevel(self, xlsx_path, sheet_name, dup_row_indices, col_indices):
        """
        执行上移操作的统一入口：优先Go方案，降级Python方案。
        """
        shifter_exe = self._get_shifter_exe_path()

        if shifter_exe and os.path.isfile(shifter_exe):
            # Go可用 → 使用Go方案（低内存、高性能）
            self._log(f"    ✓ 检测到Go处理程序: {shifter_exe}")
            self._shift_up_go(xlsx_path, sheet_name, dup_row_indices, col_indices)
        else:
            # Go不可用 → 降级到ElementTree方案
            self._log(f"    ⚠ 未找到Go处理程序，降级使用Python备用方案")
            self._shift_up_fallback(xlsx_path, sheet_name, dup_row_indices, col_indices)

    def _process_excel_with_rules(self, file_path):
        """依次执行所有已保存的规则，流式处理减少内存占用"""
        skip_header = self.skip_header_var.get()
        total_start = time.time()

        try:
            # 1. 先关闭缓存的只读workbook，拷贝源文件到临时副本
            self._close_ro_wb()
            gc.collect()

            dir_path = os.path.dirname(file_path)
            if not dir_path:
                dir_path = os.getcwd()
            base_name = os.path.basename(file_path)
            name_without_ext = os.path.splitext(base_name)[0]
            output_file = os.path.join(dir_path, f"{name_without_ext}_已去重.xlsx")

            # 检查源文件是否仍然存在（可能被移动/删除）
            if not os.path.exists(file_path):
                raise FileNotFoundError_(file_path)

            # 检查磁盘空间（粗略估算：输出文件至少需要与原文件同样大小）
            try:
                stat = os.statvfs(dir_path) if hasattr(os, 'statvfs') else None
                if stat and stat.f_bavail * stat.f_frsize < os.path.getsize(file_path) * 1.5:
                    required_mb = (os.path.getsize(file_path) * 1.5) / (1024*1024)
                    available_mb = (stat.f_bavail * stat.f_frsize) / (1024*1024)
                    raise DiskSpaceError(required_mb=required_mb, available_mb=available_mb)
            except OSError:
                pass  # Windows下可能不支持statvfs，跳过检查

            # 删除旧的输出文件
            if os.path.exists(output_file):
                try:
                    os.remove(output_file)
                    self._log(f"  已删除旧文件：{output_file}")
                except PermissionError as e:
                    raise WriteFileError(output_file, operation="删除旧文件", detail=f"文件被占用：{str(e)}")
                except Exception as e:
                    raise WriteFileError(output_file, operation="删除旧文件", detail=str(e))

            self._log("=" * 60)
            self._log("复制源文件到工作副本...")
            # 记录输出文件路径（用于异常时清理）
            self._last_output_file = output_file
            try:
                shutil.copy2(file_path, output_file)
                self._log("复制完成，开始流式处理")
            except PermissionError as e:
                raise FilePermissionError_(file_path or output_file, is_locked=True)
            except IOError as e:
                raise WriteFileError(output_file, operation="复制文件", detail=str(e))
            except MemoryError as e:
                file_mb = os.path.getsize(file_path) / (1024*1024)
                raise MemoryError_(file_size_mb=file_mb)

            # 2. 依次执行每条规则
            rule_mode = self.rule_mode_var.get()
            mode_text = "多规则联查" if rule_mode == "multi" else "单规则审查"
            self._log(f"审查模式：{mode_text}")
            
            # ====== 优先尝试Go全托管（扫描+上移一体化，内存<200MB）======
            go_success = self._process_via_go_full(output_file)
            if go_success:
                # Go全托管成功完成，跳过Python扫描流程
                elapsed = time.time() - total_start
                self._log(f"\n{'=' * 60}")
                self._log(f"✓ 全部处理完成，总耗时 {elapsed:.1f} 秒")
                self._export_log_file()
                return
            
            # Go不可用或失败 → 降级到Python原生处理（内存较高）
            self._log(f"    ⚠ Go全托管不可用，降级使用Python原生处理...")
            
            shared_seen = {} if rule_mode == "multi" else None
            all_dup_summary = {}

            for rule_idx, rule in enumerate(self.saved_rules):
                if self._cancel_flag:
                    raise InterruptedError("处理已被用户取消")

                seen = shared_seen if rule_mode == "multi" else {}
                green_selections = rule['green_selections']
                red_selections = rule['red_selections']

                involved_sheets = set(green_selections.keys()) & set(red_selections.keys())
                if not involved_sheets:
                    self._log(f"  ⚠ 规则{rule_idx + 1}的整体范围和审查列没有共同Sheet，跳过")
                    continue

                # 打印规则信息
                green_display = []
                for sn in sorted(green_selections):
                    letters = [get_column_letter(c + 1) for c in sorted(green_selections[sn])]
                    green_display.append(f"{sn}[{','.join(letters)}]")
                red_display = []
                for sn in sorted(red_selections):
                    letters = [get_column_letter(c + 1) for c in sorted(red_selections[sn])]
                    red_display.append(f"{sn}[{','.join(letters)}]")

                self._log(f"\n{'─' * 50}")
                self._log(f"▶ 规则{rule_idx + 1}：")
                self._log(f"  整体范围：{', '.join(green_display)}")
                self._log(f"  审查列：{', '.join(red_display)}")
                self._log(f"{'─' * 50}")

                scan_label = "全局共享" if rule_mode == "multi" else "规则独立"
                self._log(f"\n  ── 第一遍：流式扫描去重（{scan_label}） ──")

                # ====== 深度内存优化：消除中间存储 ======
                # ① 直接按Sheet分组（跳过 all_duplicates 中间列表）
                #    节省：~200字节/重复项 × N个重复项
                # ② 用dedup_key本身作为counter key（跳过str()转换）
                #    节省：每个重复项少一份字符串拷贝
                # ③ seen的value用整数编码代替tuple
                #    节省：(sheet,row) tuple → 单个整数，每条减少~80字节
                dup_by_sheet = {}       # {sheet_name: [row_idx, ...]} 直接分组
                dup_counter = Counter() # key=dedup_key(元组), 不转字符串
                summary_overflow_count = 0
                # Sheet名称→索引映射，用于整数编码
                _sheet_list = sorted(involved_sheets)
                _sheet_idx_map = {sn: i for i, sn in enumerate(_sheet_list)}

                # 第一遍：流式只读扫描，找出重复行号（零中间存储）
                try:
                    wb_scan = load_workbook(filename=output_file, read_only=True)
                except InvalidFileException as e:
                    raise FileCorruptError(output_file, detail=f"工作副本无法打开：{str(e)}")
                except PermissionError as e:
                    raise FilePermissionError_(output_file, is_locked=True)
                except MemoryError as e:
                    file_mb = os.path.getsize(file_path) / (1024*1024) if os.path.exists(file_path) else 0
                    raise MemoryError_(file_size_mb=file_mb)

                scanned_rows = 0

                # 合并统计+扫描为一次遍历
                for target_sheet in sorted(involved_sheets):
                    if self._cancel_flag:
                        wb_scan.close()
                        raise InterruptedError("处理已被用户取消")
                    if target_sheet not in wb_scan.sheetnames:
                        continue

                    red_cols = sorted(red_selections[target_sheet])
                    self._log(f"    扫描Sheet「{target_sheet}」审查列：{[get_column_letter(c + 1) for c in red_cols]}")

                    ws = wb_scan[target_sheet]
                    row_idx = 0
                    for row in ws.iter_rows(values_only=True):
                        self._pause_event.wait()
                        if self._cancel_flag:
                            wb_scan.close()
                            raise InterruptedError("处理已被用户取消")

                        if skip_header and row_idx == 0:
                            row_idx += 1
                            continue
                        scanned_rows += 1
                        if scanned_rows % 10000 == 0:
                            gc.collect()
                            self._log(f"    扫描进度：已扫描 {scanned_rows} 行")

                        # 避免list(row)分配：直接用元组索引
                        skip_row = False
                        if not row or len(row) <= max(red_cols, default=-1):
                            row_idx += 1
                            continue
                        key_parts = []
                        for rc in red_cols:
                            cell_value = row[rc]
                            raw = str(cell_value) if cell_value is not None else ""
                            cleaned = normalize_value(raw)
                            if not cleaned:
                                skip_row = True
                                break
                            key_parts.append(cleaned)
                        if skip_row:
                            row_idx += 1
                            continue

                        dedup_key = tuple(key_parts)

                        if dedup_key in seen:
                            # 直接追加到目标Sheet的列表（跳过中间all_duplicates）
                            dup_by_sheet.setdefault(target_sheet, []).append(row_idx)
                            # 用元组本身作为key，不生成字符串拷贝
                            dup_counter[dedup_key] += 1
                        else:
                            # 整数编码：高16位=Sheet索引，低48位=row_idx（支持百万级行号）
                            sidx = _sheet_idx_map.get(target_sheet, 0)
                            seen[dedup_key] = (sidx << 48) | row_idx
                        row_idx += 1

                wb_scan.close()
                del wb_scan
                gc.collect()

                # ====== 立即释放 seen 字典（最大内存消耗源） ======
                total_dups = sum(len(v) for v in dup_by_sheet.values())
                self._log(f"    扫描完成，共发现 {total_dups} 个重复项")
                # seen不再需要了！释放它（可能节省数百MB到数GB）
                del seen
                seen = {}
                gc.collect()

                if total_dups == 0:
                    self._log(f"    规则{rule_idx + 1}未发现重复项，跳过")
                    continue

                # 日志输出（限制数量避免UI卡顿）
                self._log(f"    重复值详情（共 {len(dup_counter)} 种重复值）：")
                shown = 0
                for dk, count in dup_counter.most_common(50):
                    display = str(dk) if len(str(dk)) <= 30 else str(dk)[:27] + "..."
                    self._log(f"      「{display}」出现 {count + 1} 次（重复 {count} 项）")
                    shown += 1
                if len(dup_counter) > shown:
                    self._log(f"      ... 还有 {len(dup_counter) - shown} 种重复值未显示")

                # 汇总统计（用row_idx占位替代content字符串，减少内存）
                for sn, row_list in dup_by_sheet.items():
                    for ridx in row_list[:min(len(row_list), MAX_SUMMARY_ENTRIES // len(dup_by_sheet) + 1)]:
                        key = (rule_idx, sn, ridx)
                        if len(all_dup_summary) < MAX_SUMMARY_ENTRIES:
                            all_dup_summary[key] = all_dup_summary.get(key, 0) + 1
                        else:
                            summary_overflow_count += 1
                if summary_overflow_count > 0:
                    self._log(f"    ⚠ 汇总统计已满（上限{MAX_SUMMARY_ENTRIES}条），{summary_overflow_count}条低频记录已省略")

                # 【内存优化】释放计数器（日志已输出完毕）
                del dup_counter
                gc.collect()

                # 第二遍：底层XML操作执行上移（零DOM膨胀）
                # 核心原理：.xlsx本质是ZIP包，每个Sheet是XML文件
                # 直接用 zipfile + ElementTree 操作XML，内存占用降低 90%+
                try:
                    for target_sheet in sorted(dup_by_sheet):
                        if self._cancel_flag:
                            raise InterruptedError("处理已被用户取消")

                        green_cols = sorted(green_selections[target_sheet])
                        sheet_dups = dup_by_sheet[target_sheet]

                        self._log(f"\n    ── 第二遍：底层上移Sheet「{target_sheet}」{len(sheet_dups)} 项 ──")

                        # 按行号从大到小排序（避免行号偏移）
                        sheet_dups.sort(reverse=True)

                        # 调用底层XML操作（不经过openpyxl DOM层）
                        self._shift_up_lowlevel(
                            output_file, target_sheet,
                            sheet_dups, green_cols
                        )

                        self._log(f"    [{target_sheet}] 上移完成，处理 {len(sheet_dups)} 项")

                        # 释放该Sheet的重复项数据
                        dup_by_sheet[target_sheet] = []
                        gc.collect()

                    # 释放所有分组数据
                    del dup_by_sheet
                    gc.collect()

                except InvalidFileException as e:
                    raise FileCorruptError(output_file, detail=f"工作副本无法以写模式打开：{str(e)}")
                except PermissionError as e:
                    raise FilePermissionError_(output_file, is_locked=True)
                except MemoryError as e:
                    file_mb = os.path.getsize(file_path) / (1024*1024) if os.path.exists(file_path) else 0
                    raise MemoryError_(file_size_mb=file_mb)
                except IOError as e:
                    raise WriteFileError(output_file, operation="保存", detail=f"文件IO错误：{str(e)}")
                except OSError as e:
                    if "no space" in str(e).lower() or "空间不足" in str(e):
                        required_mb = os.path.getsize(file_path) / (1024*1024) * 2
                        raise DiskSpaceError(required_mb=required_mb, available_mb=0)
                    else:
                        raise WriteFileError(output_file, operation="保存", detail=str(e))
                except Exception as e:
                    raise WriteFileError(output_file, operation="保存", detail=f"{type(e).__name__}: {str(e)}")

            # 3. 完成
            total_elapsed = time.time() - total_start
            total_dup_items = sum(all_dup_summary.values())
            total_dup_kinds = len(all_dup_summary)

            self._log("=" * 60)
            self._log("全部处理完成！")
            self._log(f"共执行 {len(self.saved_rules)} 条规则")
            self._log(f"共发现 {total_dup_kinds} 种重复值，处理 {total_dup_items} 个重复项")
            self._log(f"总耗时：{total_elapsed:.2f} 秒")
            self._log(f"输出文件路径：{output_file}")
            self._export_log_file()

            if all_dup_summary:
                self._log(f"\n{'─' * 50}")
                self._log("重复值汇总（按Sheet统计）：")
                # 按 Sheet 聚合统计
                sheet_stats = {}
                for (r_idx, sn, _), count in all_dup_summary.items():
                    sheet_key = f"规则{r_idx + 1}[{sn}]"
                    sheet_stats[sheet_key] = sheet_stats.get(sheet_key, 0) + count
                for sk, cnt in sorted(sheet_stats.items(), key=lambda x: -x[1])[:50]:
                    self._log(f"  {sk}：{cnt} 项")
                if len(sheet_stats) > 50:
                    self._log(f"  ... 还有 {len(sheet_stats) - 50} 个Sheet未显示")

            all_dup_summary.clear()
            gc.collect()

            # 构建弹窗消息
            result_msg = f"全部处理完成！\n\n"
            result_msg += f"共执行 {len(self.saved_rules)} 条规则\n"
            result_msg += f"共发现 {total_dup_kinds} 种重复值，处理 {total_dup_items} 个重复项\n"
            for i, rule in enumerate(self.saved_rules):
                gp = []
                for sn in sorted(rule['green_selections']):
                    letters = [get_column_letter(c + 1) for c in sorted(rule['green_selections'][sn])]
                    gp.append(f"{sn}[{','.join(letters)}]")
                rp = []
                for sn in sorted(rule['red_selections']):
                    letters = [get_column_letter(c + 1) for c in sorted(rule['red_selections'][sn])]
                    rp.append(f"{sn}[{','.join(letters)}]")
                result_msg += f"\n  规则{i + 1}：范围[{', '.join(gp)}]，审查[{', '.join(rp)}]"
            result_msg += f"\n\n输出文件：{output_file}\n"
            result_msg += f"总耗时：{total_elapsed:.2f} 秒"
            messagebox.showinfo("成功", result_msg)
            self._stop_timer(total_elapsed)

        except InterruptedError as e:
            self._export_log_file()
            # E011 - 用户主动取消
            cancel_err = UserCancelledError()
            self._safe_log_error(cancel_err)
            self._log(f"  已处理耗时：{time.time() - total_start:.2f} 秒")
            self._log("  已处理的数据未保存。")
            # 清理取消后残留的输出副本
            try:
                if 'output_file' in dir() and os.path.exists(output_file):
                    os.remove(output_file)
                    self._log(f"  已清理输出副本：{output_file}")
            except Exception:
                pass
            self._reset_timer()

        # ---- 捕获所有自定义业务异常 ----
        except ExcelCleanerError as e:
            self._export_log_file()  # 异常时也导出日志
            self._safe_log_error(e)
            # 清理残留的输出副本（仅当文件已创建时）
            if hasattr(self, '_last_output_file') and self._last_output_file:
                try:
                    if os.path.exists(self._last_output_file):
                        os.remove(self._last_output_file)
                        self._log(f"  已清理输出副本：{self._last_output_file}")
                except Exception:
                    pass
            messagebox.showerror(f"错误 {e.error_code}", e.get_user_message())
            self._reset_timer()

        # ---- 捕获系统级/未知异常（最终兜底） ----
        except MemoryError as e:
            self._export_log_file()
            file_mb = os.path.getsize(file_path) / (1024*1024) if os.path.exists(file_path) else 0
            biz_err = MemoryError_(file_size_mb=file_mb)
            self._safe_log_error(biz_err, detail=str(e))
            messagebox.showerror(f"错误 {biz_err.error_code}", biz_err.get_user_message())
            self._reset_timer()

        except PermissionError as e:
            biz_err, _ = self._classify_error(e, context="文件操作", file_path=file_path)
            self._safe_log_error(biz_err, detail=str(e))
            messagebox.showerror(f"错误 {biz_err.error_code}", biz_err.get_user_message())
            self._reset_timer()

        except OSError as e:
            if "no space" in str(e).lower():
                biz_err = DiskSpaceError(required_mb=0, available_mb=0)
            else:
                biz_err, _ = self._classify_error(e, context="系统操作", file_path=file_path)
            self._safe_log_error(biz_err, detail=str(e))
            messagebox.showerror(f"错误 {biz_err.error_code}", biz_err.get_user_message())
            self._reset_timer()

        except Exception as e:
            self._export_log_file()
            # 兜底：任何未预料到的异常
            biz_err, detail = self._classify_error(e, context="处理Excel", file_path=file_path)
            self._safe_log_error(biz_err, detail=detail)
            # 尝试清理残留的输出文件
            try:
                if 'output_file' in dir() and output_file and os.path.exists(output_file):
                    os.remove(output_file)
                    self._log(f"  已清理输出副本：{output_file}")
            except Exception:
                pass
            messagebox.showerror(f"错误 {biz_err.error_code}", biz_err.get_user_message())
            self._reset_timer()

        finally:
            self._reset_ui()

    def _lock_config_ui(self):
        """处理期间禁用左侧配置区域的所有控件（底部三按钮和右侧日志区保持可用）"""
        disabled = ("disabled",)
        for widget in [
            getattr(self, '_select_btn', None),      # 选择文件按钮
            getattr(self, '_file_entry', None),       # 文件路径输入框
            getattr(self, 'skip_header_switch', None),  # 跳过表头开关
            getattr(self, 'rule_mode_segbtn', None),   # 审查模式分段按钮
            getattr(self, 'rule_mode_help_btn', None), # 帮助图标
            getattr(self, 'green_mode_btn', None),     # 选择整体范围按钮
            getattr(self, 'red_mode_btn', None),       # 需要审查的列按钮
            getattr(self, 'save_rule_btn', None),      # 保存规则按钮
            getattr(self, 'clear_rules_btn', None),    # 清空列表按钮
        ]:
            if widget is not None:
                try:
                    widget.configure(state="disabled")
                except Exception:
                    pass
        # 禁用规则列表中的编辑/删除按钮（动态创建的）
        for btn in getattr(self, '_rule_action_buttons', []):
            try:
                btn.configure(state="disabled")
            except Exception:
                pass

    def _unlock_config_ui(self):
        """处理完成后恢复左侧配置区域的所有控件"""
        for widget in [
            getattr(self, '_select_btn', None),
            getattr(self, '_file_entry', None),
            getattr(self, 'skip_header_switch', None),
            getattr(self, 'rule_mode_segbtn', None),
            getattr(self, 'rule_mode_help_btn', None),
            getattr(self, 'green_mode_btn', None),
            getattr(self, 'red_mode_btn', None),
            getattr(self, 'save_rule_btn', None),
            getattr(self, 'clear_rules_btn', None),
        ]:
            if widget is not None:
                try:
                    widget.configure(state="normal")
                except Exception:
                    pass
        # 恢复规则列表中的编辑/删除按钮
        for btn in getattr(self, '_rule_action_buttons', []):
            try:
                btn.configure(state="normal")
            except Exception:
                pass

    def _reset_ui(self):
        """重置UI状态"""
        self.process_btn.configure(state="normal", text="🚀 开始处理")
        self.pause_btn.configure(state="disabled", text="⏸ 暂停", fg_color="#F57F17", hover_color="#F9A825")
        self.cancel_btn.configure(state="disabled")
        self._unlock_config_ui()  # 恢复左侧配置区域


def main():
    """主函数"""
    root = ctk.CTk()
    app = ExcelDeduplicationTool(root)
    root.mainloop()


if __name__ == "__main__":
    main()
